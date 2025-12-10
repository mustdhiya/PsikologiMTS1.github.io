from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Avg
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import io
import logging
import re
import json

from .models import Student
from .forms import StudentForm
from students.models import Student, StudentAchievement
from .models import Student, RMIBResult, StudentAchievement, AchievementType


# Setup logging
logger = logging.getLogger(__name__)

class IsStaffMixin(UserPassesTestMixin):
    """Mixin to check if user is staff or admin"""
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

@login_required
def student_dashboard(request):
    """
    Dashboard utama untuk student
    Redirect ke dashboard jika staff, atau tampilkan dashboard student
    """
    # Jika staff/admin, redirect ke admin area
    if request.user.is_staff or request.user.is_superuser:
        return redirect('students:list')
    
    # Ambil data student dari user
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, 'Data siswa tidak ditemukan. Hubungi administrator.')
        return redirect('accounts:login')
    
    # Ambil data RMIB
    rmib_result = None
    rmib_status = 'pending'
    top_categories = []
    
    try:
        rmib_result = RMIBResult.objects.get(student=student)
        rmib_status = rmib_result.status
        
        # ✅ FIXED: Get top 3 categories berdasarkan field yang ADA
        if rmib_result.status == 'completed':
            # Cek field mana yang ada di model RMIBResult
            if hasattr(rmib_result, 'levels') and rmib_result.levels:
                # levels adalah JSONField dengan ranking {category: rank}
                # Sort berdasarkan ranking (nilai terkecil = prioritas tertinggi)
                sorted_categories = sorted(
                    rmib_result.levels.items(), 
                    key=lambda x: x[1]  # Sort by rank value
                )[:3]
                
                # Map category key ke data lengkap
                category_mapping = {
                    'outdoor': {'name': 'Outdoor', 'icon': '🌲', 'description': 'Aktivitas luar ruangan'},
                    'mechanical': {'name': 'Mechanical', 'icon': '⚙️', 'description': 'Mesin dan teknik'},
                    'computational': {'name': 'Computational', 'icon': '🔢', 'description': 'Perhitungan dan logika'},
                    'scientific': {'name': 'Scientific', 'icon': '🔬', 'description': 'Ilmu pengetahuan'},
                    'personal_contact': {'name': 'Personal Contact', 'icon': '🤝', 'description': 'Interaksi sosial'},
                    'aesthetic': {'name': 'Aesthetic', 'icon': '🎨', 'description': 'Seni dan keindahan'},
                    'literary': {'name': 'Literary', 'icon': '📚', 'description': 'Literasi dan bahasa'},
                    'musical': {'name': 'Musical', 'icon': '🎵', 'description': 'Musik dan nada'},
                    'social_service': {'name': 'Social Service', 'icon': '❤️', 'description': 'Pelayanan sosial'},
                    'clerical': {'name': 'Clerical', 'icon': '📋', 'description': 'Administrasi'},
                    'practical': {'name': 'Practical', 'icon': '🔧', 'description': 'Praktis dan terapan'},
                    'medical': {'name': 'Medical', 'icon': '⚕️', 'description': 'Kesehatan medis'},
                }
                
                top_categories = [
                    {
                        'key': cat_key,
                        'rank': rank,
                        **category_mapping.get(cat_key, {
                            'name': cat_key.title(),
                            'icon': '📊',
                            'description': 'Kategori RMIB'
                        })
                    }
                    for cat_key, rank in sorted_categories
                ]
            
            elif hasattr(rmib_result, 'scores') and rmib_result.scores:
                # scores adalah JSONField dengan skor {category: score}
                # Sort berdasarkan skor tertinggi
                sorted_categories = sorted(
                    rmib_result.scores.items(),
                    key=lambda x: x[1],
                    reverse=True  # Skor tertinggi dulu
                )[:3]
                
                category_mapping = {
                    'outdoor': {'name': 'Outdoor', 'icon': '🌲', 'description': 'Aktivitas luar ruangan'},
                    'mechanical': {'name': 'Mechanical', 'icon': '⚙️', 'description': 'Mesin dan teknik'},
                    'computational': {'name': 'Computational', 'icon': '🔢', 'description': 'Perhitungan dan logika'},
                    'scientific': {'name': 'Scientific', 'icon': '🔬', 'description': 'Ilmu pengetahuan'},
                    'personal_contact': {'name': 'Personal Contact', 'icon': '🤝', 'description': 'Interaksi sosial'},
                    'aesthetic': {'name': 'Aesthetic', 'icon': '🎨', 'description': 'Seni dan keindahan'},
                    'literary': {'name': 'Literary', 'icon': '📚', 'description': 'Literasi dan bahasa'},
                    'musical': {'name': 'Musical', 'icon': '🎵', 'description': 'Musik dan nada'},
                    'social_service': {'name': 'Social Service', 'icon': '❤️', 'description': 'Pelayanan sosial'},
                    'clerical': {'name': 'Clerical', 'icon': '📋', 'description': 'Administrasi'},
                    'practical': {'name': 'Practical', 'icon': '🔧', 'description': 'Praktis dan terapan'},
                    'medical': {'name': 'Medical', 'icon': '⚕️', 'description': 'Kesehatan medis'},
                }
                
                top_categories = [
                    {
                        'key': cat_key,
                        'score': score,
                        **category_mapping.get(cat_key, {
                            'name': cat_key.title(),
                            'icon': '📊',
                            'description': 'Kategori RMIB'
                        })
                    }
                    for cat_key, score in sorted_categories
                ]
                
    except RMIBResult.DoesNotExist:
        pass
    
    # Ambil achievements
    achievements = StudentAchievement.objects.filter(student=student).order_by('-year', '-points')[:5]
    total_achievement_points = sum(ach.points for ach in achievements)
    
    # Certificate requests (placeholder - sesuaikan dengan model Anda)
    certificate_requests = []
    
    context = {
        'student': student,
        'rmib_result': rmib_result,
        'rmib_status': rmib_status,
        'top_categories': top_categories,
        'achievements': achievements,
        'total_achievement_points': total_achievement_points,
        'certificate_requests': certificate_requests,
    }
    
    return render(request, 'students/student_dashboard.html', context)

class StudentListView(LoginRequiredMixin, ListView):
    """Enhanced student list view with advanced filtering and search"""
    model = Student
    template_name = 'students/list.html'
    context_object_name = 'students'
    paginate_by = 25
    
    def get_queryset(self):
        # FIX: Ganti 'prestasi' dengan 'achievements'
        queryset = Student.objects.select_related(
            'user',
            'rmib_result'  # ← Tambah ini juga untuk performance
        ).prefetch_related(
            'achievements'  # ← CHANGE: prestasi → achievements
        ).order_by('student_class', 'name')
        
        # Search functionality
        search = self.request.GET.get('search', '').strip()
        if search:
            search_terms = search.split()
            for term in search_terms:
                queryset = queryset.filter(
                    Q(name__icontains=term) | 
                    Q(nisn__icontains=term) |
                    Q(student_class__icontains=term) |
                    Q(birth_place__icontains=term)
                )
        
        # Filter by class
        class_filter = self.request.GET.get('class', '').strip()
        if class_filter:
            queryset = queryset.filter(student_class=class_filter)
            
        # Filter by status
        status_filter = self.request.GET.get('status', '').strip()
        if status_filter:
            queryset = queryset.filter(test_status=status_filter)
        
        # Filter by gender
        gender_filter = self.request.GET.get('gender', '').strip()
        if gender_filter:
            queryset = queryset.filter(gender=gender_filter)
        
        # Filter by entry year
        year_filter = self.request.GET.get('year', '').strip()
        if year_filter and year_filter.isdigit():
            queryset = queryset.filter(entry_year=int(year_filter))
        
        # Sorting
        sort = self.request.GET.get('sort', '').strip()
        if sort == 'name':
            queryset = queryset.order_by('name')
        elif sort == 'nisn':
            queryset = queryset.order_by('nisn')
        elif sort == 'class':
            queryset = queryset.order_by('student_class', 'name')
        elif sort == 'status':
            queryset = queryset.order_by('test_status', 'name')
        elif sort == 'date':
            queryset = queryset.order_by('-created_at')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Preserve search parameters
        context.update({
            'search': self.request.GET.get('search', ''),
            'class_filter': self.request.GET.get('class', ''),
            'status_filter': self.request.GET.get('status', ''),
            'gender_filter': self.request.GET.get('gender', ''),
            'year_filter': self.request.GET.get('year', ''),
            'sort': self.request.GET.get('sort', ''),
        })
        
        # Get filter options
        students = Student.objects.all()
        context.update({
            'available_classes': students.values_list('student_class', flat=True).distinct().order_by('student_class'),
            'available_years': students.values_list('entry_year', flat=True).distinct().order_by('-entry_year'),
            'status_choices': Student.STATUS_CHOICES,
            'gender_choices': Student.GENDER_CHOICES,
        })
        
        # Statistics
        context.update({
            'total_students': students.count(),
            'completed_tests': students.filter(test_status='completed').count(),
            'pending_tests': students.filter(test_status='pending').count(),
            'in_progress_tests': students.filter(test_status='in_progress').count(),
        })
        
        return context

class StudentDeleteView(LoginRequiredMixin, IsStaffMixin, DeleteView):
    """Enhanced student delete view with safety checks"""
    model = Student
    template_name = 'students/delete_confirm.html'
    context_object_name = 'student'
    success_url = reverse_lazy('students:list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        
        context.update({
            'page_title': f'Hapus Siswa - {student.name}',
            'breadcrumb_title': 'Hapus Siswa',
            'has_user_account': bool(student.user),
            'has_achievements': student.achievements.exists(),  # ← FIX: prestasi → achievements
            'achievements_count': student.achievements.count(),  # ← FIX
            'has_test_results': student.test_status == 'completed',
        })
        return context
    
    def delete(self, request, *args, **kwargs):
        student = self.get_object()
        student_name = student.name
        student_nisn = student.nisn
        
        try:
            # Break relationship first to avoid cascade issues
            user_to_delete = None
            if student.user:
                user_to_delete = student.user
                student.user = None
                student.save()
            
            # Delete student
            student.delete()
            
            # Delete user if exists
            if user_to_delete:
                try:
                    user_to_delete.delete()
                except Exception as e:
                    logger.warning(f"User delete warning: {str(e)}")
            
            messages.success(
                request, 
                f'Siswa {student_name} (NISN: {student_nisn}) berhasil dihapus dari sistem!'
            )
            logger.info(f"Student deleted: {student_name} (NISN: {student_nisn}) by {request.user.username}")
                
        except Exception as e:
            logger.error(f"Error deleting student {student_name}: {str(e)}")
            messages.error(request, f'Terjadi kesalahan saat menghapus siswa: {str(e)}')
            return redirect('students:detail', pk=self.object.pk)
        
        return redirect(self.success_url)

@require_POST
@login_required
def bulk_delete_students(request):
    """Bulk delete multiple students"""
    if not request.user.is_staff:
        return JsonResponse({
            'success': False, 
            'message': 'Tidak memiliki izin'
        }, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        
        if not student_ids:
            return JsonResponse({
                'success': False,
                'message': 'Tidak ada siswa yang dipilih'
            }, status=400)
        
        # Validate IDs
        try:
            student_ids = [int(sid) for sid in student_ids]
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'message': 'ID siswa tidak valid'
            }, status=400)
        
        students = Student.objects.filter(id__in=student_ids)
        total_count = students.count()
        
        if total_count == 0:
            return JsonResponse({
                'success': False,
                'message': 'Siswa tidak ditemukan'
            }, status=404)
        
        deleted_count = 0
        failed_names = []
        
        for student in students:
            try:
                student_name = student.name
                user_to_delete = student.user if student.user else None
                
                # Break relationship
                if student.user:
                    student.user = None
                    student.save()
                
                # Delete student
                student.delete()
                
                # Delete user
                if user_to_delete:
                    try:
                        user_to_delete.delete()
                    except Exception as e:
                        logger.warning(f"User delete failed: {e}")
                
                deleted_count += 1
                logger.info(f"Student deleted: {student_name}")
                
            except Exception as e:
                logger.error(f"Delete error: {e}")
                failed_names.append(student.name)
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count} siswa berhasil dihapus',
            'deleted_count': deleted_count,
            'failed_count': len(failed_names)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format data tidak valid'
        }, status=400)
    except Exception as e:
        logger.error(f"Bulk delete error: {e}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@require_POST
@login_required
def delete_student_ajax(request, pk):
    """AJAX delete single student"""
    if not request.user.is_staff:
        return JsonResponse({
            'success': False,
            'message': 'Tidak memiliki izin'
        }, status=403)
    
    try:
        student = get_object_or_404(Student, pk=pk)
        student_name = student.name
        student_nisn = student.nisn
        user_to_delete = student.user if student.user else None
        
        # Break relationship
        if student.user:
            student.user = None
            student.save()
        
        # Delete student
        student.delete()
        
        # Delete user
        if user_to_delete:
            try:
                user_to_delete.delete()
            except Exception as e:
                logger.warning(f"User delete failed: {e}")
        
        logger.info(f"Student deleted: {student_name} by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Siswa {student_name} berhasil dihapus!',
            'redirect_url': reverse('students:list')
        })
        
    except Http404:
        return JsonResponse({
            'success': False,
            'message': 'Siswa tidak ditemukan'
        }, status=404)
    except Exception as e:
        logger.error(f"Delete error: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

class StudentCreateView(LoginRequiredMixin, IsStaffMixin, CreateView):
    """Enhanced student creation view with validation"""
    model = Student
    form_class = StudentForm
    template_name = 'students/form.html'
    success_url = reverse_lazy('students:list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': 'Tambah Siswa Baru',
            'form_action': 'Tambah',
            'breadcrumb_title': 'Tambah Siswa',
        })
        return context
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                student = form.save()
                
                user, password = student.create_user_account()
                
                messages.success(
                    self.request, 
                    f'Siswa {student.name} berhasil ditambahkan! '
                    f'NISN: {student.nisn}, Password: {password}'
                )
                
                # Log the creation
                logger.info(f"Student created: {student.name} (NISN: {student.nisn}) by {self.request.user.username}")
                
                return super().form_valid(form)
                
        except IntegrityError:
            messages.error(self.request, 'NISN sudah ada dalam sistem!')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error creating student: {str(e)}")
            messages.error(self.request, 'Terjadi kesalahan saat menyimpan data siswa')
            return self.form_invalid(form)

class StudentDetailView(DetailView):
    model = Student
    template_name = 'students/detail.html'
    context_object_name = 'student'
    
    def get_queryset(self):
        # Prefetch dengan nama relasi yang BENAR
        return Student.objects.prefetch_related(
            'achievements',  # ← Benar! (dari related_name di StudentAchievement)
            'rmib_result'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        
        # Add RMIB data
        context['rmib_result'] = student.rmib_result if hasattr(student, 'rmib_result') else None
        context['has_rmib'] = hasattr(student, 'rmib_result') and student.rmib_result.status == 'completed'
        
        # Add achievements
        context['achievements'] = student.achievements.all().order_by('-year', '-points')
        print(f"\n=== DEBUG STUDENT DETAIL ===")
        print(f"Student ID: {student.id}")
        print(f"Student Name: {student.name}")
        print(f"Student User: {student.user}")
        print(f"Student Generated Password: {student.generated_password}")
        if student.user:
            print(f"User Username: {student.user.username}")
            print(f"User Email: {student.user.email}")
        print(f"========================\n")

        return context


class StudentUpdateView(LoginRequiredMixin, IsStaffMixin, UpdateView):
    """Enhanced student update view"""
    model = Student
    form_class = StudentForm
    template_name = 'students/form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': f'Edit Data - {self.object.name}',
            'form_action': 'Update',
            'breadcrumb_title': 'Edit Siswa',
        })
        return context
    
    def get_success_url(self):
        return reverse_lazy('students:detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                old_nisn = Student.objects.get(pk=self.object.pk).nisn
                new_nisn = form.cleaned_data['nisn']
                
                student = form.save()
                
                # Update user account if NISN changed
                if old_nisn != new_nisn and student.user:
                    student.user.username = new_nisn
                    student.user.email = f"{new_nisn}@student.mts1samarinda.id"
                    student.user.save()
                
                # Update name in user account
                if student.user:
                    name_parts = student.name.split()
                    student.user.first_name = name_parts[0]
                    student.user.last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                    student.user.save()
                
                messages.success(self.request, f'Data siswa {student.name} berhasil diperbarui!')
                logger.info(f"Student updated: {student.name} (NISN: {student.nisn}) by {self.request.user.username}")
                
                return super().form_valid(form)
                
        except IntegrityError:
            messages.error(self.request, 'NISN sudah digunakan siswa lain!')
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error updating student: {str(e)}")
            messages.error(self.request, 'Terjadi kesalahan saat menyimpan data')
            return self.form_invalid(form)

@method_decorator(csrf_protect, name='dispatch')
class BatchImportView(LoginRequiredMixin, IsStaffMixin, TemplateView):
    """Enhanced batch import with better error handling and UI"""
    template_name = 'students/batch_import.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': 'Batch Import Siswa',
            'breadcrumb_title': 'Import Data',
            'max_file_size': '5MB',
            'max_records': 1000,
            'required_columns': ['nama', 'nisn', 'kelas', 'jenis_kelamin', 'tanggal_lahir'],
        })
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle both AJAX and form submissions"""
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return self.handle_ajax_upload(request)
        else:
            return self.handle_form_upload(request)
    
    def handle_ajax_upload(self, request):
        """Handle AJAX file upload with detailed JSON response"""
        try:
            if 'csv_file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'message': 'File CSV tidak ditemukan',
                    'error_type': 'no_file'
                }, status=400)
            
            csv_file = request.FILES['csv_file']
            
            # Validate file
            validation_result = self.validate_file(csv_file)
            if not validation_result['valid']:
                return JsonResponse({
                    'success': False,
                    'message': validation_result['message'],
                    'errors': validation_result.get('errors', []),
                    'error_type': 'validation_error'
                }, status=400)
            
            # Process CSV
            result = self.process_csv_file(csv_file)
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': f"Import berhasil! {result['successful']} siswa ditambahkan.",
                    'results': {
                        'total_processed': result['total_processed'],
                        'successful': result['successful'],
                        'errors': result['errors'],
                        'duplicates': result['duplicates'],
                        'error_details': result['error_details'][:10],  # Limit to 10 errors
                    },
                    'redirect_url': reverse_lazy('students:list')
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': result['message'],
                    'errors': result.get('error_details', [])[:10],
                    'error_type': 'processing_error'
                }, status=422)
                
        except Exception as e:
            logger.error(f"CSV upload error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': 'Terjadi kesalahan internal server',
                'error_type': 'server_error'
            }, status=500)
    
    def handle_form_upload(self, request):
        """Handle traditional form upload with redirect"""
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Silakan pilih file CSV untuk diupload')
            return redirect('students:batch_import')
        
        csv_file = request.FILES['csv_file']
        
        # Validate file
        validation_result = self.validate_file(csv_file)
        if not validation_result['valid']:
            messages.error(request, validation_result['message'])
            if validation_result.get('errors'):
                for error in validation_result['errors'][:5]:  # Show max 5 errors
                    messages.warning(request, error)
            return redirect('students:batch_import')
        
        # Process CSV
        result = self.process_csv_file(csv_file)
        
        if result['success']:
            messages.success(request, f"Import berhasil! {result['successful']} siswa ditambahkan.")
            if result['errors'] > 0:
                messages.warning(request, f"{result['errors']} record gagal diproses.")
            if result['duplicates'] > 0:
                messages.info(request, f"{result['duplicates']} record duplikat diabaikan.")
        else:
            messages.error(request, result['message'])
            # Show some error details
            for error in result.get('error_details', [])[:5]:
                messages.warning(request, error)
        
        return redirect('students:list')
        
    def validate_file(self, file):
        """Comprehensive file validation with detailed error reporting"""
        try:
            if not file.name.lower().endswith('.csv'):
                return {
                    'valid': False,
                    'message': 'File harus berformat CSV (.csv)',
                    'error_code': 'invalid_extension'
                }
            
            if file.size > 5 * 1024 * 1024:
                return {
                    'valid': False,
                    'message': f'Ukuran file terlalu besar ({file.size // 1024 // 1024}MB). Maksimal 5MB',
                    'error_code': 'file_too_large'
                }
            
            if file.size == 0:
                return {
                    'valid': False,
                    'message': 'File kosong. Pilih file CSV yang berisi data',
                    'error_code': 'empty_file'
                }
            
            # Read and validate CSV structure
            file.seek(0)  # Reset file pointer
            try:
                content = file.read().decode('utf-8-sig')  # Handle BOM
            except UnicodeDecodeError:
                try:
                    file.seek(0)
                    content = file.read().decode('latin-1')
                except UnicodeDecodeError:
                    return {
                        'valid': False,
                        'message': 'File tidak dapat dibaca. Pastikan encoding UTF-8 atau Latin-1',
                        'error_code': 'encoding_error'
                    }
            
            file.seek(0)  # Reset again for later use
            
            if not content.strip():
                return {
                    'valid': False,
                    'message': 'File CSV kosong atau hanya berisi whitespace',
                    'error_code': 'empty_content'
                }
            
            lines = content.strip().split('\n')
            if len(lines) < 2:
                return {
                    'valid': False,
                    'message': 'File CSV harus memiliki minimal header dan 1 baris data',
                    'error_code': 'insufficient_data'
                }
            
            # Auto-detect delimiter FIRST
            header_line = lines[0].strip()
            delimiter = self.detect_delimiter(header_line)
            
            # Parse header with detected delimiter
            header_cols = []
            if delimiter:
                # Split and clean header columns
                raw_cols = header_line.split(delimiter)
                header_cols = [col.strip().lower() for col in raw_cols if col.strip()]
            
            # Expected columns
            expected_columns = ['nama', 'nisn', 'kelas', 'jenis_kelamin', 'tanggal_lahir']
            
            # FIXED: Check if all required columns are present
            missing_cols = []
            for expected_col in expected_columns:
                if not any(self.normalize_column_name(header_col) == self.normalize_column_name(expected_col) 
                        for header_col in header_cols):
                    missing_cols.append(expected_col)
            
            if missing_cols:
                return {
                    'valid': False,
                    'message': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_cols)}',
                    'errors': [
                        f'Header ditemukan: {", ".join(header_cols)}',
                        f'Header yang diperlukan: {", ".join(expected_columns)}',
                        f'Pastikan header persis sama dengan format template'
                    ],
                    'error_code': 'missing_columns'
                }
            
            extra_cols = []
            for header_col in header_cols:
                if not any(self.normalize_column_name(header_col) == self.normalize_column_name(expected_col) 
                        for expected_col in expected_columns):
                    extra_cols.append(header_col)
            
            data_lines = len(lines) - 1  # Exclude header
            if data_lines > 1000:
                return {
                    'valid': False,
                    'message': f'Terlalu banyak record ({data_lines}). Maksimal 1000 siswa per upload',
                    'error_code': 'too_many_records'
                }
            
            # Validate a few sample rows
            sample_errors = self.validate_sample_rows(lines[1:6], delimiter, header_cols)
            if sample_errors:
                return {
                    'valid': False,
                    'message': 'Format data tidak valid pada beberapa baris',
                    'errors': sample_errors,
                    'error_code': 'data_format_error'
                }
            
            validation_result = {
                'valid': True,
                'total_records': data_lines,
                'delimiter': delimiter,
                'header_cols': header_cols
            }
            
            # Add warnings for extra columns
            if extra_cols:
                validation_result['warnings'] = [f'Kolom tambahan akan diabaikan: {", ".join(extra_cols)}']
            
            return validation_result
            
        except Exception as e:
            logger.error(f"File validation error: {str(e)}", exc_info=True)
            return {
                'valid': False,
                'message': 'Terjadi kesalahan saat memvalidasi file',
                'error_code': 'validation_exception'
            }

    def normalize_column_name(self, col_name):
        """Normalize column name for comparison"""
        if not col_name:
            return ''
        
        # Remove spaces, underscores, and convert to lowercase
        normalized = col_name.lower().strip()
        normalized = normalized.replace('_', '').replace(' ', '')
        
        # Handle common variations
        variations = {
            'namalengkap': 'nama',
            'namasiswa': 'nama',
            'jeniskelamin': 'jenis_kelamin',
            'gender': 'jenis_kelamin',
            'tanggallahir': 'tanggal_lahir',
            'tgllahir': 'tanggal_lahir',
            'birthdate': 'tanggal_lahir',
            'class': 'kelas',
            'studentclass': 'kelas',
        }
        
        return variations.get(normalized, normalized)

    def detect_delimiter(self, header):
        """Enhanced delimiter detection"""
        if not header:
            return ','
        
        delimiters = {',': 0, ';': 0, '\t': 0, '|': 0}
        
        for delimiter in delimiters:
            delimiters[delimiter] = header.count(delimiter)
        
        # Return the delimiter with highest count (and > 0)
        best_delimiter = max(delimiters.items(), key=lambda x: x[1])
        
        if best_delimiter[1] > 0:
            return best_delimiter[0]
        
        return ','  # Default to comma

    def validate_sample_rows(self, sample_rows, delimiter, header_cols):
        """Enhanced sample row validation"""
        errors = []
        expected_columns_count = len(header_cols)
        
        for i, row in enumerate(sample_rows[:3], 1):
            if not row.strip():
                continue
            
            # FIX: Handle quoted fields properly
            import csv as csv_module
            try:
                # Use CSV reader to properly parse quoted fields
                parsed_cols = list(csv_module.reader([row], delimiter=delimiter))
                if parsed_cols:
                    cols = parsed_cols[0]
                else:
                    cols = row.split(delimiter)
            except:
                cols = row.split(delimiter)
            
            # Count non-empty columns
            actual_cols_count = len([col for col in cols if col.strip()])
            
            # FIXED: Allow slight variance in column count (for optional fields)
            if actual_cols_count < expected_columns_count - 1:  # Allow 1 missing column
                errors.append(
                    f'Baris {i + 1}: Jumlah kolom kurang '
                    f'(ditemukan {actual_cols_count}, diperlukan minimal {expected_columns_count - 1})'
                )
            
            # Basic NISN validation (assuming NISN is in position based on header)
            try:
                nisn_index = None
                for idx, header in enumerate(header_cols):
                    if self.normalize_column_name(header) == 'nisn':
                        nisn_index = idx
                        break
                
                if nisn_index is not None and len(cols) > nisn_index:
                    nisn = cols[nisn_index].strip()
                    if nisn and (not nisn.isdigit() or len(nisn) != 10):
                        errors.append(f'Baris {i + 1}: Format NISN tidak valid ({nisn})')
            except (IndexError, ValueError):
                pass  # Skip if can't validate NISN
        
        return errors[:3] if errors else []  # Return empty list if no errors


    def process_csv_file(self, file):
        """Enhanced CSV processing with better column mapping"""
        result = {
            'success': False,
            'total_processed': 0,
            'successful': 0,
            'errors': 0,
            'duplicates': 0,
            'error_details': [],
            'message': ''
        }
        
        try:
            file.seek(0)
            content = file.read()
            
            # Try different encodings
            for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                try:
                    content_str = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                result['message'] = 'Tidak dapat membaca file dengan encoding yang didukung'
                return result
            
            lines = content_str.strip().split('\n')
            if not lines:
                result['message'] = 'File kosong'
                return result
            
            delimiter = self.detect_delimiter(lines[0])
            csv_reader = csv.DictReader(io.StringIO(content_str), delimiter=delimiter)
            
            if not csv_reader.fieldnames:
                result['message'] = 'File CSV tidak memiliki header'
                return result
            
            # Normalize dan map fieldnames
            field_name_mapping = {}
            normalized_fieldnames = []
            for original_name in csv_reader.fieldnames:
                normalized = self.normalize_column_name(original_name)
                field_name_mapping[normalized] = original_name.strip()
                normalized_fieldnames.append(normalized)
            
            # Verify we have all required fields
            required_fields = ['nama', 'nisn', 'kelas', 'jeniskelamin', 'tanggallahir']
            missing_fields = [field for field in required_fields if field not in normalized_fieldnames]
            
            if missing_fields:
                result['message'] = f'Field yang diperlukan tidak ditemukan: {", ".join(missing_fields)}'
                return result
            
            processed_nisns = set()  # Track processed NISNs
            
            for row_num, row in enumerate(csv_reader, start=2):
                result['total_processed'] += 1
                
                try:
                    # Map row data using field mapping
                    mapped_row = {}
                    for normalized_field, original_field in field_name_mapping.items():
                        if normalized_field in ['nama', 'nisn', 'kelas', 'jeniskelamin', 'tanggallahir', 'tempatlahir', 'password']:
                            mapped_row[normalized_field] = row.get(original_field, '').strip()
                    
                    cleaned_data = self.clean_row_data_mapped(mapped_row, row_num)
                    
                    if not cleaned_data['valid']:
                        result['errors'] += len(cleaned_data['errors'])
                        result['error_details'].extend(cleaned_data['errors'])
                        continue
                    
                    data = cleaned_data['data']
                    
                    if data['nisn'] in processed_nisns:
                        result['duplicates'] += 1
                        result['error_details'].append(f"Baris {row_num}: NISN {data['nisn']} duplikat dalam file")
                        continue
                    
                    processed_nisns.add(data['nisn'])
                    
                    if Student.objects.filter(nisn=data['nisn']).exists():
                        result['duplicates'] += 1
                        result['error_details'].append(f"Baris {row_num}: NISN {data['nisn']} sudah ada dalam sistem")
                        continue
                    
                    # Extract custom password
                    custom_password = data.pop('custom_password', None)
                    
                    # Create student OUTSIDE atomic block
                    student = Student.objects.create(**data)
                    
                    # Try to create user account - SEPARATE transaction
                    try:
                        if not User.objects.filter(username=data['nisn']).exists():
                            if custom_password:
                                # Use custom password from CSV
                                user, password = student.create_user_account(custom_password=custom_password)
                            else:
                                # Auto-generate password
                                user, password = student.create_user_account()
                        else:
                            logger.warning(f"User account already exists for NISN: {data['nisn']}")
                    except IntegrityError as e:
                        logger.warning(f"Failed to create user for {student.name} (User duplicate) - {str(e)}")
                    except Exception as e:
                        logger.warning(f"Failed to create user account for {student.name}: {str(e)}")
                    
                    result['successful'] += 1
                    
                except IntegrityError as e:
                    result['errors'] += 1
                    error_msg = f"Baris {row_num}: Data tidak valid (kemungkinan NISN duplikat) - {str(e)}"
                    result['error_details'].append(error_msg)
                    logger.warning(f'IntegrityError: {error_msg}')
                except Exception as e:
                    result['errors'] += 1
                    error_msg = f"Baris {row_num}: {str(e)}"
                    result['error_details'].append(error_msg)
                    logger.error(f'Row processing error: {error_msg}')
            
            # Determine success status dan message
            if result['successful'] > 0:
                result['success'] = True
                result['message'] = f"Import selesai. {result['successful']} siswa berhasil ditambahkan."
                if result['errors'] > 0 or result['duplicates'] > 0:
                    result['message'] += f" ({result['errors']} error, {result['duplicates']} duplikat)."
            else:
                result['message'] = 'Tidak ada siswa yang berhasil ditambahkan.'
            
            if result['successful'] > 0:
                logger.info(f'Batch import completed: {result["successful"]} students created by {self.request.user.username}')
        
        except Exception as e:
            logger.error(f'CSV processing error: {str(e)}', exc_info=True)
            result['message'] = f'Terjadi kesalahan saat memproses file CSV: {str(e)}'
            result['error_details'].append(f'System error: {str(e)}')
        
        return result

    def clean_row_data_mapped(self, row, row_num):
        """Clean row data with mapped field names - SUPPORT PASSWORD"""
        errors = []
        
        try:
            # Extract and clean data
            nama = str(row.get('nama', '')).strip()
            nisn = str(row.get('nisn', '')).strip()
            kelas = str(row.get('kelas', '')).strip().upper()
            jenis_kelamin_raw = str(row.get('jeniskelamin', '')).strip().upper()
            tanggal_lahir = str(row.get('tanggallahir', '')).strip()
            tempat_lahir = str(row.get('tempatlahir', '')).strip()  # NEW
            password = str(row.get('password', '')).strip()  # NEW - Support custom password
            
            # Validate nama - SIMPLE VERSION (no regex)
            if not nama:
                errors.append(f"Baris {row_num}: Nama tidak boleh kosong")
            elif len(nama) < 2:
                errors.append(f"Baris {row_num}: Nama terlalu pendek (minimal 2 karakter)")
            elif len(nama) > 200:
                errors.append(f"Baris {row_num}: Nama terlalu panjang (maksimal 200 karakter)")
            
            # Validate NISN
            if not nisn:
                errors.append(f"Baris {row_num}: NISN tidak boleh kosong")
            elif not nisn.isdigit():
                errors.append(f"Baris {row_num}: NISN harus berupa angka (ditemukan: {nisn})")
            elif len(nisn) != 10:
                errors.append(f"Baris {row_num}: NISN harus 10 digit (ditemukan: {len(nisn)} digit)")
            
            # Validate kelas
            if not kelas:
                errors.append(f"Baris {row_num}: Kelas tidak boleh kosong")
            elif not re.match(r'^[7-9][A-Z]$', kelas):
                errors.append(f"Baris {row_num}: Format kelas tidak valid (harus 7A, 8B, 9C, dll). Ditemukan: {kelas}")
            
            # Validate dan mapping jenis kelamin - FIXED!
            jenis_kelamin_mapping = {
                'L': 'L', 'P': 'P',
                'LAKI': 'L', 'LAKI-LAKI': 'L', 'LELAKI': 'L', 'MALE': 'L', 'M': 'L',
                'PEREMPUAN': 'P', 'WANITA': 'P', 'FEMALE': 'P', 'F': 'P',
            }
            jenis_kelamin = jenis_kelamin_mapping.get(jenis_kelamin_raw)
            
            if not jenis_kelamin:
                errors.append(
                    f"Baris {row_num}: Jenis kelamin tidak valid "
                    f"(gunakan L/P, LAKI-LAKI, atau PEREMPUAN). Ditemukan: {jenis_kelamin_raw}"
                )
            
            # Validate dan parse tanggal lahir
            birth_date = None
            if not tanggal_lahir:
                errors.append(f"Baris {row_num}: Tanggal lahir tidak boleh kosong")
            else:
                # Try multiple date formats
                date_formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d%m%Y']
                parsed = False
                for date_format in date_formats:
                    try:
                        birth_date = datetime.strptime(tanggal_lahir, date_format).date()
                        
                        # Handle 2-digit years
                        if birth_date.year < 1950:  # Assume it's 20xx
                            birth_date = birth_date.replace(year=birth_date.year + 100)
                        
                        parsed = True
                        break
                    except ValueError:
                        continue
                
                if not parsed:
                    errors.append(f"Baris {row_num}: Format tanggal lahir tidak valid (gunakan DD/MM/YYYY)")
                else:
                    # Validate date range
                    today = datetime.now().date()
                    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                    
                    if age < 10 or age > 25:
                        errors.append(f"Baris {row_num}: Usia tidak wajar ({age} tahun)")
                    
                    if birth_date > today:
                        errors.append(f"Baris {row_num}: Tanggal lahir tidak boleh di masa depan")
            
            if errors:
                return {'valid': False, 'errors': errors}
            
            # Validate password if provided
            if password and (len(password) < 4 or len(password) > 20):
                errors.append(f"Baris {row_num}: Password harus 4-20 karakter (ditemukan: {len(password)})")
            
            if errors:
                return {'valid': False, 'errors': errors}
            
            # Return cleaned data
            return {
                'valid': True,
                'data': {
                    'name': nama.title().strip(),
                    'nisn': nisn,
                    'student_class': kelas,
                    'gender': jenis_kelamin,  # Now mapped correctly (L atau P)
                    'birth_date': birth_date,
                    'birth_place': tempat_lahir.title() if tempat_lahir else '',  # NEW
                    'custom_password': password if password else None,  # NEW - Store custom password
                    'entry_year': datetime.now().year,
                    'test_status': 'pending',
                    'phone': '',
                    'address': '',
                    'parent_phone': '',
                }
            }
            
        except Exception as e:
            return {'valid': False, 'errors': [f"Baris {row_num}: Error memproses data - {str(e)}"]}


@require_http_methods(["POST"])
@csrf_protect
def reset_student_password(request, pk):
    """Reset password for specific student (AJAX endpoint)"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        student = get_object_or_404(Student, pk=pk)
        new_password = student.reset_password()
        
        if new_password:
            logger.info(f"Password reset by {request.user.username} for student: {student.name}")
            return JsonResponse({
                'success': True,
                'message': f'Password berhasil direset untuk {student.name}',
                'new_password': new_password
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Gagal mereset password. Siswa belum memiliki akun user.'
            })
            
    except Exception as e:
        logger.error(f"Password reset error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Terjadi kesalahan saat mereset password'
        })

@require_http_methods(["POST"])
@csrf_protect 
def unlock_student_account(request, pk):
    """Unlock student account (AJAX endpoint)"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        student = get_object_or_404(Student, pk=pk)
        student.unlock_account()
        
        logger.info(f"Account unlocked by {request.user.username} for student: {student.name}")
        return JsonResponse({
            'success': True,
            'message': f'Akun {student.name} berhasil dibuka kembali'
        })
        
    except Exception as e:
        logger.error(f"Account unlock error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Terjadi kesalahan saat membuka akun'
        })

def export_students_csv(request):
    """Export students data as CSV"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki izin untuk mengekspor data')
        return redirect('students:list')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="data_siswa_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 'Tanggal Lahir', 'Tempat Lahir', 'Status Tes', 'Tahun Masuk'])
    
    students = Student.objects.all().order_by('student_class', 'name')
    for student in students:
        writer.writerow([
            student.nisn,
            student.name,
            student.student_class,
            student.get_gender_display(),
            student.birth_date.strftime('%d/%m/%Y'),
            student.birth_place,
            student.get_test_status_display(),
            student.entry_year
        ])
    
    logger.info(f"Students data exported by {request.user.username}")
    return response

def download_csv_template(request):
    """Download CSV template for batch import"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="template_import_siswa.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['nama', 'nisn', 'kelas', 'jenis_kelamin', 'tanggal_lahir'])
    writer.writerow(['Ahmad Fauzan', '1234567890', '8A', 'L', '15/03/2010'])
    writer.writerow(['Siti Nurhaliza', '0987654321', '8A', 'P', '22/07/2010'])
    
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
import json
import logging
from datetime import datetime

from students.models import Student, RMIBResult, StudentAchievement, AchievementType

logger = logging.getLogger(__name__)

# ==================== RMIB CATEGORIES ====================
RMIB_CATEGORIES = {
    'outdoor': {
        'name': 'Outdoor (Alam Terbuka)',
        'description': 'Aktivitas yang berhubungan dengan alam dan lingkungan luar',
        'icon': 'fas fa-tree',
        'color': 'green'
    },
    'mechanical': {
        'name': 'Mechanical (Mekanik)',
        'description': 'Pekerjaan dengan mesin, alat, dan teknologi',
        'icon': 'fas fa-cog',
        'color': 'blue'
    },
    'computational': {
        'name': 'Computational (Komputasi)',
        'description': 'Bekerja dengan angka, data, dan analisis',
        'icon': 'fas fa-calculator',
        'color': 'purple'
    },
    'scientific': {
        'name': 'Scientific (Sains)',
        'description': 'Penelitian, eksperimen, dan penemuan ilmiah',
        'icon': 'fas fa-flask',
        'color': 'indigo'
    },
    'personal_contact': {
        'name': 'Personal Contact (Hubungan Personal)',
        'description': 'Berinteraksi dan membantu orang lain',
        'icon': 'fas fa-handshake',
        'color': 'pink'
    },
    'aesthetic': {
        'name': 'Aesthetic (Estetika)',
        'description': 'Seni, desain, dan keindahan',
        'icon': 'fas fa-palette',
        'color': 'orange'
    },
    'literary': {
        'name': 'Literary (Sastra)',
        'description': 'Menulis, membaca, dan komunikasi verbal',
        'icon': 'fas fa-book',
        'color': 'teal'
    },
    'musical': {
        'name': 'Musical (Musik)',
        'description': 'Musik, suara, dan harmoni',
        'icon': 'fas fa-music',
        'color': 'red'
    },
    'social_service': {
        'name': 'Social Service (Pelayanan Sosial)',
        'description': 'Membantu masyarakat dan kesejahteraan sosial',
        'icon': 'fas fa-hands-helping',
        'color': 'amber'
    },
    'clerical': {
        'name': 'Clerical (Administratif)',
        'description': 'Administrasi, organisasi, dan tata kelola',
        'icon': 'fas fa-file-alt',
        'color': 'gray'
    },
    'practical': {
        'name': 'Practical (Praktis)',
        'description': 'Pekerjaan praktis dan aplikatif sehari-hari',
        'icon': 'fas fa-tools',
        'color': 'yellow'
    },
    'medical': {
        'name': 'Medical (Medis)',
        'description': 'Kesehatan dan perawatan medis',
        'icon': 'fas fa-heartbeat',
        'color': 'red'
    }
}

# ==================== RMIB TEST VIEWS ====================

@login_required
def rmib_test_interface(request, student_pk):
    """Display RMIB test interface with achievement input"""
    student = get_object_or_404(Student, pk=student_pk)
    
    if not student.can_take_test():
        messages.warning(request, 'Siswa ini sudah menyelesaikan tes RMIB.')
        return redirect('students:detail', pk=student_pk)
    
    has_progress = hasattr(student, 'rmib_result') and student.rmib_result.levels
    rmib_categories_json = json.dumps(RMIB_CATEGORIES)
    
    context = {
        'student': student,
        'rmib_categories': RMIB_CATEGORIES,
        'rmib_categories_json': rmib_categories_json,
        'has_progress': has_progress,
        'total_categories': len(RMIB_CATEGORIES),
        'current_year': datetime.now().year
    }
    
    logger.info(f"🔍 RMIB Test Interface - Student: {student.name}, Has Progress: {has_progress}")
    
    return render(request, 'students/rmib_test.html', context)


@login_required
@require_http_methods(["POST"])
def start_rmib_test(request, student_pk):
    """Start or resume RMIB test"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        
        rmib_result, created = RMIBResult.objects.get_or_create(
            student=student,
            defaults={
                'submitted_at': timezone.now(),
                'status': 'in_progress'
            }
        )
        
        if not created:
            rmib_result.status = 'in_progress'
            rmib_result.save()
        
        logger.info(f"Starting RMIB test for {student.name} - Created: {created}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tes dimulai' if created else 'Tes dilanjutkan',
            'has_progress': not created
        })
    except Exception as e:
        logger.error(f"Start test error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def save_rmib_progress(request, student_pk):
    """Save RMIB test progress (autosave)"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        data = json.loads(request.body)
        
        levels = data.get('levels', {})
        
        rmib_result, created = RMIBResult.objects.get_or_create(
            student=student,
            defaults={
                'submitted_at': timezone.now(),
                'levels': levels
            }
        )
        
        if not created:
            rmib_result.levels = levels
            rmib_result.save()
        
        logger.info(f"Saved progress for {student.name}: {len(levels)} levels")
        
        return JsonResponse({
            'success': True,
            'message': 'Progress berhasil disimpan',
            'saved_at': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Save progress error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def load_rmib_progress(request, student_pk):
    """Load saved RMIB progress"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        
        if hasattr(student, 'rmib_result') and student.rmib_result.levels:
            return JsonResponse({
                'success': True,
                'has_progress': True,
                'levels': student.rmib_result.levels,
                'saved_at': student.rmib_result.updated_at.isoformat() if student.rmib_result.updated_at else None
            })
        else:
            return JsonResponse({
                'success': True,
                'has_progress': False,
                'levels': {}
            })
            
    except Exception as e:
        logger.error(f"Load progress error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def submit_rmib_test(request, student_pk):
    """Submit RMIB test + Achievements - Integrated"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        data = json.loads(request.body)
        
        levels = data.get('levels', {})
        achievements_data = data.get('achievements', [])
        
        logger.info(f"Submitting RMIB + Achievements for {student.name}: {len(levels)} levels, {len(achievements_data)} achievements")
        
        # Validate levels - all 12 categories must have levels
        if len(levels) != 12:
            return JsonResponse({
                'success': False,
                'message': f'Semua 12 kategori harus diisi. Diterima: {len(levels)}'
            }, status=400)
        
        # Validate level range (1-12)
        for category, level in levels.items():
            try:
                level_int = int(level)
                if level_int < 1 or level_int > 12:
                    return JsonResponse({
                        'success': False,
                        'message': f'Level harus antara 1-12. Kategori {category}: {level}'
                    }, status=400)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': f'Level harus berupa angka. Kategori {category}: {level}'
                }, status=400)
        
        # Create or update RMIB result
        rmib_result, created = RMIBResult.objects.get_or_create(
            student=student,
            defaults={'submitted_at': timezone.now()}
        )
        
        # Save levels dan calculate scores
        rmib_result.levels = levels
        rmib_result.calculate_scores()
        rmib_result.submitted_at = timezone.now()
        rmib_result.status = 'completed'
        rmib_result.save()
        
        # ==================== PROCESS ACHIEVEMENTS ====================
        if achievements_data:
            for ach_data in achievements_data:
                try:
                    achievement_type = AchievementType.objects.get(pk=ach_data['achievement_type_id'])
                    
                    # Check if already exists
                    existing = StudentAchievement.objects.filter(
                        student=student,
                        achievement_type=achievement_type,
                        year=int(ach_data['year'])
                    ).first()
                    
                    if not existing:
                        achievement = StudentAchievement.objects.create(
                            student=student,
                            achievement_type=achievement_type,
                            level=ach_data['level'],
                            rank=ach_data['rank'],
                            year=int(ach_data['year']),
                            notes=ach_data.get('notes', '')
                        )
                        
                        # Auto-verify if submitted by staff
                        if request.user.is_staff:
                            achievement.verify(request.user)
                        
                        logger.info(f"Created achievement: {achievement_type.name} ({achievement.points} poin)")
                    
                except AchievementType.DoesNotExist:
                    logger.warning(f"Achievement type not found: {ach_data['achievement_type_id']}")
                except Exception as e:
                    logger.error(f"Error creating achievement: {str(e)}")
                    continue
        
        # ==================== UPDATE STUDENT STATUS ====================
        student.test_status = 'completed'
        student.test_date = timezone.now()
        student.save(update_fields=['test_status', 'test_date'])
        
        # Calculate totals
        verified_achievements = StudentAchievement.objects.filter(student=student, is_verified=True)
        total_achievement_score = sum(ach.points for ach in verified_achievements)
        test_score = rmib_result.total_score
        combined = test_score + total_achievement_score
        
        logger.info(f"✅ RMIB + Achievements submitted for {student.name}. Test: {test_score} + Achievements: {total_achievement_score} = {combined}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tes RMIB dan prestasi berhasil diselesaikan!',
            'redirect_url': f'/students/{student_pk}/rmib/result/',
            'test_score': test_score,
            'achievement_score': total_achievement_score,
            'combined_score': combined,
            'primary_interest': rmib_result.primary_interest,
            'primary_level': rmib_result.primary_level
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format data tidak valid'
        }, status=400)
    except Exception as e:
        logger.error(f"Submit error: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def rmib_result_view(request, student_pk):
    """Display RMIB results + Achievements - RANKING #1 = 60 POIN"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        
        if not hasattr(student, 'rmib_result'):
            messages.warning(request, 'Anda belum menyelesaikan tes RMIB.')
            return redirect('students:rmib_test', student_pk=student_pk)
        
        rmib_result = student.rmib_result
        
        if not rmib_result.levels:
            messages.warning(request, 'Data tes belum tersimpan dengan baik.')
            return redirect('students:rmib_test', student_pk=student_pk)
        
        # ==================== BUILD RANKING (SCORE DESCENDING) ====================
        ranking_data = []
        
        for category_key, level in rmib_result.levels.items():
            category_info = RMIB_CATEGORIES.get(category_key, {})
            level_int = int(level)
            # SCORE = (13 - LEVEL) * 5
            # Level 1 = (13-1)*5 = 60 poin (Ranking #1)
            # Level 12 = (13-12)*5 = 5 poin (Ranking #12)
            score = (13 - level_int) * 5
            
            ranking_data.append({
                'rank': 0,  # Will assign after sort
                'category_key': category_key,
                'category_name': category_info.get('name', category_key),
                'level': level_int,
                'score': score,
                'icon': category_info.get('icon', 'fas fa-circle'),
                'color': category_info.get('color', 'gray')
            })
        
        # SORT by SCORE DESCENDING (60, 55, 50... = Ranking #1, #2, #3...)
        ranking_data.sort(key=lambda x: (-x['score'], -x['level']))
        
        # Assign ranking
        for idx, item in enumerate(ranking_data, start=1):
            item['rank'] = idx
        
        logger.info(f"📊 Ranking untuk {student.name}:")
        for item in ranking_data[:3]:
            logger.info(f"   Rank #{item['rank']}: {item['category_name']} (Level {item['level']}, Score {item['score']})")
        
        # Get primary interest (rank #1 = 60 poin)
        primary_item = ranking_data[0] if ranking_data else None
        primary_interest_name = primary_item['category_name'] if primary_item else 'N/A'
        primary_level = primary_item['level'] if primary_item else 0
        primary_score = primary_item['score'] if primary_item else 0
        
        # ==================== GET ACHIEVEMENTS ====================
        achievements = StudentAchievement.objects.filter(
            student=student,
            is_verified=True
        ).select_related('achievement_type').order_by('-points', '-year')
        
        logger.info(f"Found {achievements.count()} verified achievements")
        
        # Calculate achievement contributions
        achievement_contributions = {}
        total_achievement_score = 0
        
        for achievement in achievements:
            points = achievement.points
            total_achievement_score += points
            
            if hasattr(achievement.achievement_type, 'rmib_primary') and achievement.achievement_type.rmib_primary:
                primary_cat = str(achievement.achievement_type.rmib_primary).lower()
                achievement_contributions[primary_cat] = achievement_contributions.get(primary_cat, 0) + points
            
            if hasattr(achievement.achievement_type, 'rmib_secondary') and achievement.achievement_type.rmib_secondary:
                secondary_cat = str(achievement.achievement_type.rmib_secondary).lower()
                secondary_points = points // 2 if points % 2 == 0 else (points + 1) // 2
                achievement_contributions[secondary_cat] = achievement_contributions.get(secondary_cat, 0) + secondary_points
        
        # ==================== GET COMBINED SCORES ====================
        combined_scores = {}
        for item in ranking_data:
            cat_key = item['category_key']
            cat_name = item['category_name']
            test_score = item['score']
            ach_score = achievement_contributions.get(cat_key, 0)
            total = test_score + ach_score
            
            combined_scores[cat_name] = {
                'category': cat_name,
                'test_score': test_score,
                'achievement_score': ach_score,
                'total_score': total
            }
        
        test_score_total = sum(item['score'] for item in ranking_data)
        combined_score_total = test_score_total + total_achievement_score
        
        # ==================== BUILD CONTEXT ====================
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'categories': ranking_data,
            'test_score': test_score_total,
            'achievement_score': total_achievement_score,
            'combined_score': combined_score_total,
            'primary_interest': primary_interest_name,
            'primary_level': primary_level,
            'primary_score': primary_score,
            'achievements': achievements,
            'achievement_contributions': achievement_contributions,
            'combined_scores': combined_scores
        }
        
        logger.info(f"✅ Rank #1: {primary_interest_name} ({primary_score} poin) - Level {primary_level}")
        
        return render(request, 'students/rmib_result.html', context)
    
    except Exception as e:
        logger.error(f"❌ RMIB result view error: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('core:dashboard')


# ==================== RMIB EDIT FUNCTIONS ====================

@login_required
@require_http_methods(["GET"])
def rmib_edit_confirmation(request, student_pk):
    """Show confirmation before editing RMIB"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        
        if not hasattr(student, 'rmib_result'):
            messages.error(request, 'Tidak ada hasil RMIB untuk diedit.')
            return redirect('students:detail', pk=student_pk)
        
        rmib_result = student.rmib_result
        
        # FIX: Set submitted_at if empty
        if rmib_result.submitted_at is None:
            rmib_result.submitted_at = timezone.now()
            rmib_result.save(update_fields=['submitted_at'])
            logger.info(f"Auto-set submitted_at for {student.name}")
        
        # Check if levels exist
        if not rmib_result.levels:
            messages.warning(request, 'Tes belum diselesaikan.')
            return redirect('students:rmib_test', student_pk=student_pk)
        
        # Ensure status is correct
        if rmib_result.status not in ['completed', 'edited']:
            rmib_result.status = 'completed'
            rmib_result.save(update_fields=['status'])
        
        context = {
            'student': student,
            'rmib_result': rmib_result
        }
        
        return render(request, 'students/rmib_edit_confirmation.html', context)
    
    except Exception as e:
        logger.error(f"Edit confirmation error: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('students:detail', pk=student_pk)


@login_required
@require_http_methods(["GET", "POST"])
def rmib_restart_test(request, student_pk):
    """Restart RMIB test - reset data for editing"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        
        if hasattr(student, 'rmib_result'):
            rmib_result = student.rmib_result
            rmib_result.levels = {}
            rmib_result.status = 'in_progress'
            rmib_result.save()
            
            # Also update student status
            student.test_status = 'in_progress'
            student.save(update_fields=['test_status'])
            
            logger.info(f"✅ RMIB restarted for {student.name}")
            
            if request.method == 'POST':
                return JsonResponse({
                    'success': True,
                    'message': 'RMIB siap untuk diedit'
                })
            else:
                messages.success(request, 'RMIB direset. Silakan mulai dari awal.')
                return redirect('students:rmib_test', student_pk=student_pk)
        
        return redirect('students:detail', pk=student_pk)
    
    except Exception as e:
        logger.error(f"Restart error: {str(e)}")
        if request.method == 'POST':
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        else:
            messages.error(request, 'Gagal mereset RMIB')
            return redirect('students:detail', pk=student_pk)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def submit_rmib_test_edited(request, student_pk):
    """Submit edited RMIB test"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        data = json.loads(request.body)
        levels = data.get('levels', {})
        
        logger.info(f"Submitting edited RMIB for {student.name}")
        
        if len(levels) != 12:
            return JsonResponse({'success': False, 'message': 'Semua 12 kategori harus diisi'}, status=400)
        
        for category, level in levels.items():
            try:
                level_int = int(level)
                if level_int < 1 or level_int > 12:
                    return JsonResponse({'success': False, 'message': 'Level harus antara 1-12'}, status=400)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'message': 'Level harus berupa angka'}, status=400)
        
        rmib_result = get_object_or_404(RMIBResult, student=student)
        rmib_result.levels = levels
        rmib_result.calculate_scores()
        rmib_result.status = 'edited'
        rmib_result.edited_at = timezone.now()
        rmib_result.save()
        
        student.test_status = 'completed'
        student.test_date = timezone.now()
        student.save(update_fields=['test_status', 'test_date'])
        
        logger.info(f"✅ Edited RMIB submitted for {student.name}")
        
        return JsonResponse({
            'success': True,
            'message': 'Hasil RMIB berhasil diperbarui!',
            'redirect_url': f'/students/{student_pk}/rmib/result/',
            'combined_score': rmib_result.total_score
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Format data tidak valid'}, status=400)
    except Exception as e:
        logger.error(f"Submit edited error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def rmib_cancel_edit(request, student_pk):
    """Cancel edit dan kembali ke hasil"""
    try:
        return redirect('students:rmib_result', student_pk=student_pk)
    except Exception as e:
        logger.error(f"Cancel edit error: {str(e)}")
        return redirect('students:detail', pk=student_pk)


@login_required
def export_rmib_pdf(request, student_pk):
    """Export RMIB result as PDF"""
    try:
        student = get_object_or_404(Student, pk=student_pk)
        rmib_result = get_object_or_404(RMIBResult, student=student)
        
        html_content = f"""
        <html>
        <head>
            <title>Hasil RMIB - {student.name}</title>
            <style>
                body {{ font-family: Arial; margin: 20px; }}
                h1 {{ color: #333; }}
                .info {{ background: #f0f0f0; padding: 10px; margin: 10px 0; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
            </style>
        </head>
        <body>
            <h1>Hasil Tes RMIB</h1>
            <div class="info">
                <p><strong>Nama:</strong> {student.name}</p>
                <p><strong>NISN:</strong> {student.nisn}</p>
                <p><strong>Kelas:</strong> {student.student_class}</p>
                <p><strong>Total Skor:</strong> {rmib_result.total_score}</p>
            </div>
            <h2>Ranking Kategori</h2>
            <table>
                <tr>
                    <th>Kategori</th>
                    <th>Level</th>
                    <th>Skor</th>
                </tr>
        """
        
        for cat_key, level in rmib_result.levels.items():
            score = int(level) * 5
            html_content += f"<tr><td>{cat_key}</td><td>{level}</td><td>{score}</td></tr>"
        
        html_content += """
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="RMIB_{student.name}.html"'
        return response
    
    except Exception as e:
        logger.error(f"Export PDF error: {str(e)}")
        messages.error(request, 'Gagal export hasil')
        return redirect('students:rmib_result', student_pk=student_pk)


# ==================== API ENDPOINTS ====================
@login_required
@require_http_methods(["GET"])
def api_achievement_types(request):
    """Get all active achievement types with RMIB categories - for form dropdown"""
    try:
        types = AchievementType.objects.filter(is_active=True).order_by('category', 'name').values(
            'id', 'name', 'category', 'description', 'rmib_primary', 'rmib_secondary'
        )
        
        result = []
        for t in types:
            result.append({
                'id': int(t['id']),
                'name': str(t['name']),
                'category': str(t['category']),
                'description': str(t['description'] or ''),
                'rmib_primary': str(t['rmib_primary']) if t['rmib_primary'] else None,
                'rmib_secondary': str(t['rmib_secondary']) if t['rmib_secondary'] else None
            })
        
        logger.info(f"✅ API Response: {len(result)} achievement types")
        
        return JsonResponse(result, safe=False, status=200)
        
    except Exception as e:
        logger.error(f"❌ API Error: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': str(e),
            'status': 'error'
        }, status=500)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import timedelta
import json
import logging
from accounts.views import StudentRequiredMixin
from students.models import Student, RMIBResult, CertificateRequest


logger = logging.getLogger(__name__)

# RMIB Categories mapping - Sesuaikan dengan kategori Anda
RMIB_CATEGORIES = {
    'pemimpin': {'name': 'Kepemimpinan', 'icon': 'fas fa-crown'},
    'sosial': {'name': 'Sosial', 'icon': 'fas fa-people-group'},
    'perlindungan': {'name': 'Perlindungan Sosial', 'icon': 'fas fa-shield'},
    'bisnis': {'name': 'Bisnis', 'icon': 'fas fa-briefcase'},
    'sastra': {'name': 'Sastra', 'icon': 'fas fa-book'},
    'seni': {'name': 'Seni', 'icon': 'fas fa-palette'},
    'musik': {'name': 'Musik', 'icon': 'fas fa-music'},
    'layanan': {'name': 'Layanan Umum', 'icon': 'fas fa-handshake'},
    'teknik': {'name': 'Teknik', 'icon': 'fas fa-gears'},
    'pertanian': {'name': 'Pertanian', 'icon': 'fas fa-leaf'},
    'matematika': {'name': 'Matematika', 'icon': 'fas fa-calculator'},
    'alam': {'name': 'Ilmu Alam', 'icon': 'fas fa-flask'},
}


# ==================== STUDENT CERTIFICATE PAGE ====================

@login_required
def student_certificate_page(request):
    """Halaman certificate untuk siswa"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, 'Data siswa tidak ditemukan.')
        return redirect('accounts:login')
    
    # Cek apakah sudah ada hasil RMIB
    try:
        rmib_result = RMIBResult.objects.get(student=student)
        # Ada hasil RMIB - tampilkan certificate dashboard
        ranking_data = build_ranking_data(rmib_result)
        
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'ranking_data': ranking_data,
            'primary_interest': ranking_data[0] if ranking_data else None,
            'page_title': 'Sertifikat RMIB',
            'has_result': True
        }
        return render(request, 'students/certificate_view.html', context)
    except RMIBResult.DoesNotExist:
        # Belum ada hasil RMIB - tampilkan halaman waiting
        context = {
            'student': student,
            'page_title': 'Menunggu Hasil RMIB',
            'has_result': False
        }
        return render(request, 'students/waiting_rmib.html', context)


# ==================== VIEW CERTIFICATE ====================
@login_required
def view_certificate(request):
    """Lihat sertifikat detail - tanpa parameter"""
    try:
        student = Student.objects.get(user=request.user)
        rmib_result = RMIBResult.objects.get(student=student)
        
        ranking_data = build_ranking_data(rmib_result)
        primary_interest = ranking_data[0] if ranking_data else None
        
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'ranking_data': ranking_data,
            'primary_interest': primary_interest,
            'request_date': rmib_result.submitted_at,
        }
        
        logger.info(f"Certificate viewed: {student.name}")
        return render(request, 'students/certificate_view.html', context)
        
    except (Student.DoesNotExist, RMIBResult.DoesNotExist):
        messages.error(request, 'Data sertifikat tidak ditemukan.')
        return redirect('students:certificate_page')
    except Exception as e:
        logger.error(f"View certificate error: {str(e)}")
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('students:certificate_page')


@login_required
def view_parent_report(request):
    """Lihat laporan orang tua - FIXED"""
    try:
        student = Student.objects.get(user=request.user)
        rmib_result = RMIBResult.objects.get(student=student)
        
        # Build ranking data
        ranking_data = build_ranking_data(rmib_result)
        primary_interest = ranking_data[0] if ranking_data else {
            'category_name': 'Belum Tersedia',
            'level': 0,
            'score': 0,
            'rank': 0
        }
        
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'ranking_data': ranking_data,
            'primary_interest': primary_interest,
            'request_date': rmib_result.submitted_at or timezone.now(),
        }
        
        logger.info(f"Parent report context: ranking_data={len(ranking_data)}, primary={primary_interest}")
        return render(request, 'students/parent_report_view.html', context)
        
    except (Student.DoesNotExist, RMIBResult.DoesNotExist):
        messages.error(request, 'Data laporan tidak ditemukan.')
        return redirect('students:certificate_page')
    except Exception as e:
        logger.error(f"View parent report error: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('students:certificate_page')


@login_required
def view_summary(request):
    """Lihat ringkasan - FIXED"""
    try:
        student = Student.objects.get(user=request.user)
        rmib_result = RMIBResult.objects.get(student=student)
        
        # Build ranking data
        ranking_data = build_ranking_data(rmib_result)
        primary_interest = ranking_data[0] if ranking_data else {
            'category_name': 'N/A',
            'level': 0,
            'score': 0,
            'rank': 0
        }
        
        # Calculate percentage
        persentase = int((rmib_result.total_score / 720) * 100) if rmib_result.total_score else 0
        
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'ranking_data': ranking_data,
            'primary_interest': primary_interest,
            'persentase': persentase,
            'request_date': rmib_result.submitted_at or timezone.now(),
        }
        
        logger.info(f"Summary context: ranking_data={len(ranking_data)}, persentase={persentase}")
        return render(request, 'students/summary_view.html', context)
        
    except (Student.DoesNotExist, RMIBResult.DoesNotExist):
        messages.error(request, 'Data ringkasan tidak ditemukan.')
        return redirect('students:certificate_page')
    except Exception as e:
        logger.error(f"View summary error: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('students:certificate_page')


@login_required
def view_certificate(request):
    """Lihat sertifikat - FIXED"""
    try:
        student = Student.objects.get(user=request.user)
        rmib_result = RMIBResult.objects.get(student=student)
        
        # Build ranking data
        ranking_data = build_ranking_data(rmib_result)
        primary_interest = ranking_data[0] if ranking_data else {
            'category_name': 'N/A',
            'level': 0,
            'score': 0,
            'rank': 1
        }
        
        context = {
            'student': student,
            'rmib_result': rmib_result,
            'ranking_data': ranking_data,
            'primary_interest': primary_interest,
            'request_date': rmib_result.submitted_at or timezone.now(),
            'request_id': 1,
        }
        
        logger.info(f"Certificate context: ranking_data={len(ranking_data)}")
        return render(request, 'students/certificate_view.html', context)
        
    except (Student.DoesNotExist, RMIBResult.DoesNotExist):
        messages.error(request, 'Data sertifikat tidak ditemukan.')
        return redirect('students:certificate_page')
    except Exception as e:
        logger.error(f"View certificate error: {str(e)}", exc_info=True)
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('students:certificate_page')


# ==================== REQUEST CERTIFICATE ====================

@login_required
@require_http_methods(["POST"])
def request_certificate(request, template_type):
    """Student request certificate"""
    try:
        # Validate template type
        valid_types = ['certificate', 'summary', 'parent']
        if template_type not in valid_types:
            return JsonResponse({
                'success': False, 
                'message': 'Template tidak valid'
            }, status=400)
        
        # Get student
        student = Student.objects.get(user=request.user)
        
        # Check if student has RMIB result
        try:
            rmib_result = RMIBResult.objects.get(student=student)
        except RMIBResult.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Anda belum menyelesaikan tes RMIB'
            }, status=400)
        
        # Check if already has pending request
        pending = CertificateRequest.objects.filter(
            student=student,
            status='pending'
        ).first()
        
        if pending:
            return JsonResponse({
                'success': False,
                'message': f'Anda masih memiliki permintaan sertifikat yang sedang diproses'
            }, status=400)
        
        # Create new request
        cert_request = CertificateRequest.objects.create(
            student=student,
            template_type=template_type,
            status='generated'  # Auto-generate untuk student
        )
        
        # Set generated time
        cert_request.generated_at = timezone.now()
        cert_request.save()
        
        logger.info(f"Certificate request created: {student.name} - {template_type}")
        
        # Redirect ke view yang sesuai
        redirect_url = f'/students/certificate/{template_type}/'
        if template_type == 'certificate':
            redirect_url = f'/students/certificate/view/{cert_request.id}/'
        elif template_type == 'summary':
            redirect_url = '/students/certificate/summary/'
        elif template_type == 'parent':
            redirect_url = '/students/certificate/parent-report/'
        
        return JsonResponse({
            'success': True,
            'message': f'Sertifikat {template_type} siap ditampilkan',
            'request_id': cert_request.id,
            'redirect_url': redirect_url
        })
    
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': 'Data siswa tidak ditemukan'
        }, status=400)
    except Exception as e:
        logger.error(f"Request certificate error: {str(e)}")
        return JsonResponse({
            'success': False, 
            'message': str(e)
        }, status=500)


# ==================== DOWNLOAD CERTIFICATE ====================

@login_required
@require_http_methods(["POST"])
def download_certificate_pdf(request, request_id):
    """Download certificate as PDF"""
    try:
        student = Student.objects.get(user=request.user)
        cert_request = get_object_or_404(CertificateRequest, id=request_id, student=student)
        
        if cert_request.status not in ['generated', 'downloaded']:
            return JsonResponse({
                'success': False, 
                'message': 'Sertifikat belum siap untuk diunduh'
            }, status=400)
        
        # Mark as downloaded
        cert_request.downloaded_at = timezone.now()
        cert_request.status = 'downloaded'
        cert_request.save()
        
        # Generate PDF name
        pdf_filename = f"RMIB_{student.name}_{cert_request.template_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        logger.info(f"Certificate downloaded: {student.name} - {pdf_filename}")
        
        return JsonResponse({
            'success': True,
            'message': 'Sertifikat siap diunduh',
            'filename': pdf_filename
        })
    
    except Exception as e:
        logger.error(f"Download certificate error: {str(e)}")
        return JsonResponse({
            'success': False, 
            'message': str(e)
        }, status=500)


# ==================== CANCEL CERTIFICATE ====================

@login_required
@require_http_methods(["POST"])
def cancel_certificate_request(request, request_id):
    """Cancel pending certificate request"""
    try:
        student = Student.objects.get(user=request.user)
        cert_request = get_object_or_404(CertificateRequest, id=request_id, student=student)
        
        if cert_request.status != 'pending':
            return JsonResponse({
                'success': False, 
                'message': 'Hanya permintaan yang pending dapat dibatalkan'
            }, status=400)
        
        # Delete request
        template_type = cert_request.template_type
        cert_request.delete()
        
        logger.info(f"Certificate request cancelled: {student.name} - {template_type}")
        
        return JsonResponse({
            'success': True,
            'message': f'Permintaan sertifikat {template_type} telah dibatalkan'
        })
    
    except Exception as e:
        logger.error(f"Cancel certificate error: {str(e)}")
        return JsonResponse({
            'success': False, 
            'message': str(e)
        }, status=500)


# ==================== GET CERTIFICATE STATUS ====================

@login_required
@require_http_methods(["GET"])
def get_certificate_status(request, request_id):
    """Get certificate request status (for AJAX polling)"""
    try:
        student = Student.objects.get(user=request.user)
        cert_request = get_object_or_404(CertificateRequest, id=request_id, student=student)
        
        return JsonResponse({
            'success': True,
            'status': cert_request.status,
            'template_type': cert_request.template_type,
            'requested_at': cert_request.requested_at.isoformat(),
            'generated_at': cert_request.generated_at.isoformat() if cert_request.generated_at else None,
            'downloaded_at': cert_request.downloaded_at.isoformat() if cert_request.downloaded_at else None,
        })
    
    except Exception as e:
        logger.error(f"Get status error: {str(e)}")
        return JsonResponse({
            'success': False, 
            'message': str(e)
        }, status=500)


# ==================== HELPER FUNCTIONS ====================

RMIB_CATEGORIES = {
    'outdoor': {'name': 'Outdoor', 'icon': 'fas fa-tree'},
    'mechanical': {'name': 'Mekanik', 'icon': 'fas fa-wrench'},
    'computational': {'name': 'Komputasional', 'icon': 'fas fa-laptop'},
    'scientific': {'name': 'Ilmiah', 'icon': 'fas fa-flask'},
    'personal_contact': {'name': 'Kontak Personal', 'icon': 'fas fa-handshake'},
    'aesthetic': {'name': 'Estetika', 'icon': 'fas fa-palette'},
    'literary': {'name': 'Sastra', 'icon': 'fas fa-book'},
    'musical': {'name': 'Musik', 'icon': 'fas fa-music'},
    'social_service': {'name': 'Layanan Sosial', 'icon': 'fas fa-heart'},
    'clerical': {'name': 'Klerikal', 'icon': 'fas fa-file-alt'},
    'practical': {'name': 'Praktis', 'icon': 'fas fa-tools'},
    'medical': {'name': 'Medis', 'icon': 'fas fa-stethoscope'},
}


def build_ranking_data(rmib_result):
    """
    Build ranking data dari RMIB result - FINAL FIXED VERSION
    Handles list of tuples from get_ranking_summary() dan levels dict
    """
    ranking_data = []
    
    try:
        logger.info(f"Building ranking data for: {rmib_result.student.name}")
        
        # METHOD 1: Gunakan get_ranking_summary() yang return list of tuples
        if hasattr(rmib_result, 'get_ranking_summary'):
            rankings = rmib_result.get_ranking_summary()
            logger.info(f"Rankings type: {type(rankings)}, length: {len(rankings) if rankings else 0}")
            
            # Jika return list of tuples: [('medical', 12), ('practical', 11), ...]
            if isinstance(rankings, list) and rankings:
                for idx, item in enumerate(rankings, 1):
                    if isinstance(item, tuple) and len(item) == 2:
                        category_key, level = item
                    elif isinstance(item, dict):
                        category_key = item.get('category', 'unknown')
                        level = item.get('level', 0)
                    else:
                        continue
                    
                    category_info = RMIB_CATEGORIES.get(category_key, {})
                    level_int = int(level) if level else 0
                    score = level_int * 5  # Skor = Level * 5
                    
                    ranking_data.append({
                        'rank': idx,
                        'category_key': category_key,
                        'category_name': category_info.get('name', category_key.replace('_', ' ').title()),
                        'level': level_int,
                        'score': score,
                        'icon': category_info.get('icon', 'fas fa-circle'),
                    })
                    
                    logger.info(f"Added ranking: {category_key} - Level {level_int}, Score {score}")
        
        # METHOD 2: Fallback ke levels dict jika get_ranking_summary tidak ada atau kosong
        if not ranking_data and hasattr(rmib_result, 'levels') and rmib_result.levels:
            levels_data = rmib_result.levels
            logger.info(f"Fallback to levels: {levels_data}")
            
            # Sort by level descending
            sorted_levels = sorted(levels_data.items(), key=lambda x: -x[1])
            
            for idx, (category_key, level) in enumerate(sorted_levels, 1):
                category_info = RMIB_CATEGORIES.get(category_key, {})
                level_int = int(level) if level else 0
                score = level_int * 5
                
                ranking_data.append({
                    'rank': idx,
                    'category_key': category_key,
                    'category_name': category_info.get('name', category_key.replace('_', ' ').title()),
                    'level': level_int,
                    'score': score,
                    'icon': category_info.get('icon', 'fas fa-circle'),
                })
        
        # Jika masih kosong, log warning
        if not ranking_data:
            logger.warning(f"No ranking data found for {rmib_result.student.name}")
        
        logger.info(f"Successfully built {len(ranking_data)} ranking items")
        return ranking_data
    
    except Exception as e:
        logger.error(f"Error building ranking data: {str(e)}", exc_info=True)
        return []



def get_primary_interest(ranking_data):
    """Get primary (first) interest dari ranking data"""
    return ranking_data[0] if ranking_data else None


def get_interest_description(category_key):
    """Get description untuk kategori interest"""
    descriptions = {
        'pemimpin': 'Anda memiliki potensi kepemimpinan yang kuat dan cocok untuk posisi managerial.',
        'sosial': 'Anda memiliki minat sosial yang tinggi dan cocok untuk pekerjaan yang melibatkan orang lain.',
        'seni': 'Anda memiliki kreativitas seni yang tinggi dan cocok untuk bidang seni dan desain.',
        'teknik': 'Anda memiliki minat teknis yang tinggi dan cocok untuk bidang engineering dan teknologi.',
        'bisnis': 'Anda memiliki kemampuan bisnis yang baik dan cocok untuk entrepreneur atau manajemen bisnis.',
    }
    return descriptions.get(category_key, 'Anda memiliki potensi yang unik pada bidang ini.')

# ==================== RMIB BATCH IMPORT (OPTIMIZED) ====================

@method_decorator(csrf_protect, name='dispatch')
class RMIBBatchImportView(LoginRequiredMixin, IsStaffMixin, TemplateView):
    """Batch Import RMIB Results - OPTIMIZED VERSION"""
    template_name = 'students/rmib_batch_import.html'
    
    # RMIB Column Mapping (CSV header → model field)
    RMIB_COLUMN_MAP = {
        'Out': 'outdoor',
        'Me': 'mechanical',
        'COMP': 'computational',
        'Sci': 'scientific',
        'Prs': 'personal_contact',
        'Aesth': 'aesthetic',
        'Lit': 'literary',
        'Mus': 'musical',
        'S.S': 'social_service',
        'Cler': 'clerical',
        'Prac': 'practical',
        'Med': 'medical'
    }
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': 'Batch Import Hasil RMIB',
            'breadcrumb_title': 'Import RMIB Batch',
            'max_file_size': '10MB',
            'max_records': 2000,
            'rmib_categories': list(self.RMIB_COLUMN_MAP.keys()),
            'required_columns': [
                'NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 'Tanggal Lahir',
                'Out', 'Me', 'COMP', 'Sci', 'Prs', 'Aesth', 'Lit', 'Mus', 'S.S', 'Cler', 'Prac', 'Med'
            ],
        })
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle upload - support AJAX & form"""
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return self.handle_ajax_upload(request)
        return self.handle_form_upload(request)
    
    def handle_ajax_upload(self, request):
        """AJAX upload with detailed JSON response"""
        try:
            if 'csv_file' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'message': 'File CSV tidak ditemukan',
                    'error_code': 'no_file'
                }, status=400)
            
            csv_file = request.FILES['csv_file']
            
            # Validate
            validation = self.validate_file(csv_file)
            if not validation['valid']:
                return JsonResponse({
                    'success': False,
                    'message': validation['message'],
                    'errors': validation.get('errors', []),
                    'error_code': 'validation_failed'
                }, status=400)
            
            # Process
            result = self.process_csv(csv_file)
            
            if result['success']:
                return JsonResponse({
                    'success': True,
                    'message': result['message'],
                    'results': {
                        'total_processed': result['total_processed'],
                        'students_created': result['students_created'],
                        'students_updated': result['students_updated'],
                        'rmib_created': result['rmib_created'],
                        'rmib_updated': result['rmib_updated'],
                        'rmib_skipped': result['rmib_skipped'],
                        'errors': result['errors'],
                        'error_details': result['error_details'][:15],
                    },
                    'redirect_url': reverse_lazy('students:list')
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': result['message'],
                    'errors': result.get('error_details', [])[:15],
                    'error_code': 'processing_failed'
                }, status=422)
                
        except Exception as e:
            logger.error(f"AJAX upload error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan server: {str(e)}',
                'error_code': 'server_error'
            }, status=500)
    
    def handle_form_upload(self, request):
        """Traditional form upload with redirect"""
        if 'csv_file' not in request.FILES:
            messages.error(request, 'Silakan pilih file CSV')
            return redirect('students:rmib_batch_import')
        
        csv_file = request.FILES['csv_file']
        
        validation = self.validate_file(csv_file)
        if not validation['valid']:
            messages.error(request, validation['message'])
            for error in validation.get('errors', [])[:5]:
                messages.warning(request, error)
            return redirect('students:rmib_batch_import')
        
        result = self.process_csv(csv_file)
        
        if result['success']:
            messages.success(request, result['message'])
            if result['rmib_skipped'] > 0:
                messages.info(request, f"{result['rmib_skipped']} siswa sudah selesai RMIB (diabaikan).")
            if result['errors'] > 0:
                messages.warning(request, f"{result['errors']} baris gagal diproses.")
        else:
            messages.error(request, result['message'])
            for error in result.get('error_details', [])[:5]:
                messages.warning(request, error)
        
        return redirect('students:list')
    
    def validate_file(self, file):
        """Comprehensive file validation"""
        try:
            # Extension check
            if not file.name.lower().endswith('.csv'):
                return {
                    'valid': False,
                    'message': 'File harus berformat CSV (.csv)',
                    'error_code': 'invalid_extension'
                }
            
            # Size check
            if file.size > 10 * 1024 * 1024:
                return {
                    'valid': False,
                    'message': f'File terlalu besar ({file.size // 1024 // 1024}MB). Maksimal 10MB',
                    'error_code': 'file_too_large'
                }
            
            if file.size == 0:
                return {
                    'valid': False,
                    'message': 'File kosong',
                    'error_code': 'empty_file'
                }
            
            # Read content
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                file.seek(0)
                try:
                    content = file.read().decode('latin-1')
                except UnicodeDecodeError:
                    return {
                        'valid': False,
                        'message': 'Encoding file tidak didukung. Gunakan UTF-8',
                        'error_code': 'encoding_error'
                    }
            
            file.seek(0)
            
            if not content.strip():
                return {
                    'valid': False,
                    'message': 'File kosong atau hanya berisi whitespace',
                    'error_code': 'empty_content'
                }
            
            lines = content.strip().split('\n')
            if len(lines) < 2:
                return {
                    'valid': False,
                    'message': 'File harus memiliki header dan minimal 1 baris data',
                    'error_code': 'insufficient_rows'
                }
            
            # Check header
            header = lines[0].strip()
            delimiter = self.detect_delimiter(header)
            
            # Required columns
            required_student = ['NISN', 'Nama', 'Kelas']
            required_rmib = list(self.RMIB_COLUMN_MAP.keys())
            
            missing = []
            for col in required_student + required_rmib:
                if col not in header:
                    missing.append(col)
            
            if missing:
                return {
                    'valid': False,
                    'message': f'Kolom tidak lengkap: {", ".join(missing)}',
                    'errors': [
                        f'Header ditemukan: {header}',
                        f'Kolom yang diperlukan: {", ".join(required_student + required_rmib)}'
                    ],
                    'error_code': 'missing_columns'
                }
            
            # Check record count
            data_rows = len(lines) - 1
            if data_rows > 2000:
                return {
                    'valid': False,
                    'message': f'Terlalu banyak record ({data_rows}). Maksimal 2000',
                    'error_code': 'too_many_records'
                }
            
            return {
                'valid': True,
                'total_records': data_rows,
                'delimiter': delimiter
            }
            
        except Exception as e:
            logger.error(f"File validation error: {str(e)}")
            return {
                'valid': False,
                'message': 'Terjadi kesalahan validasi',
                'error_code': 'validation_exception'
            }
    
    def detect_delimiter(self, header):
        """Auto-detect CSV delimiter"""
        delimiters = {',': 0, ';': 0, '\t': 0, '|': 0}
        
        for delimiter in delimiters:
            delimiters[delimiter] = header.count(delimiter)
        
        best = max(delimiters.items(), key=lambda x: x[1])
        return best[0] if best[1] > 0 else ','
    
    def process_csv(self, file):
        """Process CSV file - OPTIMIZED with bulk operations"""
        result = {
            'success': False,
            'total_processed': 0,
            'students_created': 0,
            'students_updated': 0,
            'rmib_created': 0,
            'rmib_updated': 0,
            'rmib_skipped': 0,
            'errors': 0,
            'error_details': [],
            'message': ''
        }
        
        try:
            file.seek(0)
            content = file.read()
            
            # Decode
            for encoding in ['utf-8-sig', 'utf-8', 'latin-1']:
                try:
                    content_str = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                result['message'] = 'Tidak dapat membaca file'
                return result
            
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(content_str))
            
            # Track processed NISN
            processed_nisn = set()
            
            # Batch lists (for potential bulk_create optimization)
            students_to_create = []
            
            for row_num, row in enumerate(csv_reader, start=2):
                result['total_processed'] += 1
                
                try:
                    # ==================== EXTRACT DATA ====================
                    nisn = str(row.get('NISN', '')).strip()
                    
                    if nisn.startswith("'"):
                        nisn = nisn[1:]
                    nisn = ''.join(filter(str.isdigit, nisn)).strip()
                    
                    nama = str(row.get('Nama', '')).strip()
                    kelas = str(row.get('Kelas', '')).strip().upper()
                    jenis_kelamin = str(row.get('Jenis Kelamin', '')).strip()
                    tanggal_lahir_str = str(row.get('Tanggal Lahir', '')).strip()
                    tempat_lahir = str(row.get('Tempat Lahir', '')).strip()
                    tahun_masuk_str = str(row.get('Tahun Masuk', '')).strip()
                    
                    # ==================== VALIDATION ====================
                    # Validate NISN (sekarang sudah bersih)
                    if not nisn or len(nisn) != 10:
                        result['errors'] += 1
                        result['error_details'].append(
                            f"Baris {row_num}: NISN tidak valid (harus 10 digit, dapat: '{nisn}')"
                        )
                        continue
                    
                    # Check duplicate in file
                    if nisn in processed_nisn:
                        result['errors'] += 1
                        result['error_details'].append(f"Baris {row_num}: NISN {nisn} duplikat dalam file")
                        continue
                    
                    processed_nisn.add(nisn)
                    
                    # Validate nama
                    if not nama or len(nama) < 2:
                        result['errors'] += 1
                        result['error_details'].append(f"Baris {row_num}: Nama tidak valid")
                        continue
                    
                    # Parse gender
                    gender = 'L' if 'laki' in jenis_kelamin.lower() or jenis_kelamin.upper() == 'L' else 'P'
                    
                    # Parse date
                    tanggal_lahir = self.parse_date(tanggal_lahir_str)
                    
                    # Parse year
                    try:
                        tahun_masuk = int(tahun_masuk_str) if tahun_masuk_str.isdigit() else datetime.now().year
                    except:
                        tahun_masuk = datetime.now().year
                    
                    # ==================== STUDENT DATA ====================
                    student_data = {
                        'name': nama,
                        'student_class': kelas,
                        'gender': gender,
                        'birth_date': tanggal_lahir,
                        'birth_place': tempat_lahir,
                        'entry_year': tahun_masuk,
                    }
                    
                    # Get or create student
                    student, created = Student.objects.get_or_create(
                        nisn=nisn,
                        defaults=student_data
                    )
                    
                    if created:
                        result['students_created'] += 1
                        # Create user account
                        try:
                            student.create_user_account()
                        except Exception as e:
                            logger.warning(f"Failed to create user for {nisn}: {str(e)}")
                    else:
                        # Update existing student (optional fields)
                        updated = False
                        for field, value in student_data.items():
                            if value and getattr(student, field) != value:
                                setattr(student, field, value)
                                updated = True
                        
                        if updated:
                            student.save()
                            result['students_updated'] += 1
                    
                    # ==================== RMIB DATA ====================
                    rmib_rankings = {}
                    has_rmib_data = False
                    
                    # Extract RMIB rankings
                    for csv_col, category_key in self.RMIB_COLUMN_MAP.items():
                        ranking_str = str(row.get(csv_col, '')).strip()
                        
                        if ranking_str:
                            try:
                                ranking = int(ranking_str)
                                if 1 <= ranking <= 12:
                                    rmib_rankings[category_key] = ranking
                                    has_rmib_data = True
                                else:
                                    result['error_details'].append(
                                        f"Baris {row_num}: Ranking {csv_col} di luar range 1-12 ({ranking})"
                                    )
                            except ValueError:
                                result['error_details'].append(
                                    f"Baris {row_num}: Ranking {csv_col} bukan angka ({ranking_str})"
                                )
                    
                    # Process RMIB if data exists
                    if has_rmib_data:
                        # Check if already completed (SKIP if completed)
                        if hasattr(student, 'rmib_result') and student.rmib_result.status == 'completed':
                            result['rmib_skipped'] += 1
                            logger.debug(f"Skipped NISN {nisn} - already completed")
                            continue
                        
                        # Determine status based on completeness
                        if len(rmib_rankings) == 12:
                            rmib_status = 'completed'
                            test_status = 'completed'
                        else:
                            rmib_status = 'in_progress'
                            test_status = 'in_progress'
                            result['error_details'].append(
                                f"Baris {row_num}: Data RMIB tidak lengkap ({len(rmib_rankings)}/12) - status in_progress"
                            )
                        
                        # Create or update RMIB result
                        rmib_result, rmib_created = RMIBResult.objects.update_or_create(
                            student=student,
                            defaults={
                                'levels': rmib_rankings,
                                'submitted_at': timezone.now(),
                                'status': rmib_status
                            }
                        )
                        
                        # Calculate scores
                        rmib_result.calculate_scores()
                        rmib_result.save()
                        
                        # Update student status
                        student.test_status = test_status
                        if rmib_status == 'completed':
                            student.test_date = timezone.now()
                        student.save()
                        
                        if rmib_created:
                            result['rmib_created'] += 1
                        else:
                            result['rmib_updated'] += 1
                        
                        logger.info(f"RMIB {rmib_status} for {nama} (NISN: {nisn})")
                
                except Exception as e:
                    result['errors'] += 1
                    error_msg = f"Baris {row_num}: {str(e)}"
                    result['error_details'].append(error_msg)
                    logger.error(f"Row processing error: {error_msg}")
            
            # ==================== FINALIZE RESULT ====================
            if result['rmib_created'] > 0 or result['students_created'] > 0:
                result['success'] = True
                parts = []
                if result['students_created'] > 0:
                    parts.append(f"{result['students_created']} siswa baru")
                if result['students_updated'] > 0:
                    parts.append(f"{result['students_updated']} siswa di-update")
                if result['rmib_created'] > 0:
                    parts.append(f"{result['rmib_created']} RMIB baru")
                if result['rmib_updated'] > 0:
                    parts.append(f"{result['rmib_updated']} RMIB di-update")
                
                result['message'] = f"Import selesai: {', '.join(parts)}."
                
                logger.info(
                    f"RMIB Batch Import completed by {self.request.user.username}: "
                    f"{result['rmib_created']} created, {result['rmib_updated']} updated"
                )
            else:
                result['message'] = "Tidak ada data yang berhasil di-import."
            
        except Exception as e:
            logger.error(f"CSV processing error: {str(e)}", exc_info=True)
            result['message'] = f'Terjadi kesalahan: {str(e)}'
            result['error_details'].append(f"System error: {str(e)}")
        
        return result
    
    def parse_date(self, date_str):
        """Parse date from various formats"""
        if not date_str:
            return None
        
        formats = [
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d.%m.%Y',
            '%Y.%m.%d'
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        logger.warning(f"Failed to parse date: {date_str}")
        return None


# ==================== TEMPLATE DOWNLOAD ====================

@login_required
def download_rmib_template(request):
    """Download template CSV untuk RMIB batch import"""
    if not request.user.is_staff:
        return HttpResponse('Unauthorized', status=403)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template_rmib_batch_import.csv"'
    response.write('\ufeff')  # UTF-8 BOM for Excel
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow([
        'NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 'Tanggal Lahir', 'Tempat Lahir', 
        'Status Tes', 'Tahun Masuk',
        'Out', 'Me', 'COMP', 'Sci', 'Prs', 'Aesth', 'Lit', 'Mus', 'S.S', 'Cler', 'Prac', 'Med'
    ])
    
    # Example 1: Completed
    writer.writerow([
        '1234567890', 'Ahmad Ramadhan', '7A', 'Laki-laki', '2008-05-15', 'Jakarta',
        'completed', '2021',
        '1', '3', '2', '4', '5', '6', '7', '8', '9', '10', '11', '12'
    ])
    
    # Example 2: Completed dengan ranking berbeda
    writer.writerow([
        '1234567891', 'Siti Nurhaliza', '7A', 'Perempuan', '2008-08-22', 'Bandung',
        'completed', '2021',
        '2', '4', '1', '3', '6', '5', '8', '7', '9', '11', '10', '12'
    ])
    
    # Example 3: In Progress (partial data)
    writer.writerow([
        '1234567892', 'Budi Santoso', '7B', 'Laki-laki', '2008-03-10', 'Surabaya',
        'in_progress', '2021',
        '1', '2', '', '', '5', '', '', '', '', '', '', ''
    ])
    
    # Example 4: Pending (no RMIB data)
    writer.writerow([
        '1234567893', 'Dewi Kusuma', '7B', 'Perempuan', '2008-11-30', 'Medan',
        'pending', '2021',
        '', '', '', '', '', '', '', '', '', '', '', ''
    ])
    
    logger.info(f"RMIB template downloaded by {request.user.username}")
    
    return response
class RMIBTestView(LoginRequiredMixin, TemplateView):
    template_name = 'students/rmib_test.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # ✅ GET student dari URL parameter
        student_pk = self.kwargs.get('pk')
        student = get_object_or_404(Student, pk=student_pk)
        
        # RMIB categories definition
        rmib_categories = {
            'outdoor': {'name': 'Outdoor', 'icon': 'fa-tree', 'desc': 'Aktivitas luar ruangan', 'color': '#10b981'},
            'mechanical': {'name': 'Mechanical', 'icon': 'fa-cog', 'desc': 'Mesin dan teknik', 'color': '#f59e0b'},
            'computational': {'name': 'Computational', 'icon': 'fa-calculator', 'desc': 'Perhitungan', 'color': '#3b82f6'},
            'scientific': {'name': 'Scientific', 'icon': 'fa-flask', 'desc': 'Ilmu pengetahuan', 'color': '#8b5cf6'},
            'personal_contact': {'name': 'Personal Contact', 'icon': 'fa-handshake', 'desc': 'Interaksi sosial', 'color': '#ec4899'},
            'aesthetic': {'name': 'Aesthetic', 'icon': 'fa-palette', 'desc': 'Seni', 'color': '#f43f5e'},
            'literary': {'name': 'Literary', 'icon': 'fa-book', 'desc': 'Literasi', 'color': '#06b6d4'},
            'musical': {'name': 'Musical', 'icon': 'fa-music', 'desc': 'Musik', 'color': '#14b8a6'},
            'social_service': {'name': 'Social Service', 'icon': 'fa-heart', 'desc': 'Pelayanan sosial', 'color': '#ef4444'},
            'clerical': {'name': 'Clerical', 'icon': 'fa-clipboard', 'desc': 'Administrasi', 'color': '#84cc16'},
            'practical': {'name': 'Practical', 'icon': 'fa-tools', 'desc': 'Praktis', 'color': '#f97316'},
            'medical': {'name': 'Medical', 'icon': 'fa-heartbeat', 'desc': 'Kesehatan medis', 'color': '#0ea5e9'},
        }
        
        # Check for existing progress
        has_progress = False
        existing_levels = {}
        existing_scores = {}
        
        try:
            rmib_result = RMIBResult.objects.get(student=student)
            if rmib_result.levels:
                has_progress = True
                existing_levels = rmib_result.levels
                existing_scores = rmib_result.scores or {}
        except RMIBResult.DoesNotExist:
            pass
        
        # Get existing achievements
        achievements = StudentAchievement.objects.filter(student=student).select_related('achievement_type')
        achievements_data = []
        for ach in achievements:
            achievements_data.append({
                'id': ach.id,
                'type_name': ach.achievement_type.name,
                'type_id': ach.achievement_type.id,
                'level': ach.level,
                'rank': ach.rank,
                'year': ach.year,
                'points': ach.points,
                'notes': ach.notes or '',
                'rmib_category': ach.achievement_type.rmib_category,
            })
        
        context.update({
            'student': student,  # ← PENTING!
            'rmib_categories_json': json.dumps(rmib_categories),
            'has_progress': has_progress,
            'existing_levels': json.dumps(existing_levels),
            'existing_scores': json.dumps(existing_scores),
            'existing_achievements': json.dumps(achievements_data),
        })
        
        return context
@require_POST
@login_required
def rmib_start_test(request, student_id):
    """Start RMIB test for student"""
    try:
        # ✅ FIX: Tambah error handling untuk student not found
        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            logger.error(f"Start test error: No Student with ID {student_id}")
            return JsonResponse({
                'success': False,
                'message': f'Siswa dengan ID {student_id} tidak ditemukan'
            }, status=404)
        
        # Check if user has permission (only own student or staff)
        if not request.user.is_staff and (not hasattr(request.user, 'student') or request.user.student.id != student.id):
            return JsonResponse({
                'success': False,
                'message': 'Anda tidak memiliki akses ke tes ini'
            }, status=403)
        
        # Get or create RMIB result
        rmib_result, created = RMIBResult.objects.get_or_create(
            student=student,
            defaults={
                'status': 'in_progress',
                'levels': {},
                'scores': {}
            }
        )
        
        if not created and rmib_result.status == 'completed':
            return JsonResponse({
                'success': False,
                'message': 'Tes RMIB sudah selesai. Hubungi admin untuk reset.'
            })
        
        # Update status
        rmib_result.status = 'in_progress'
        rmib_result.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Tes dimulai',
            'has_progress': bool(rmib_result.levels)
        })
        
    except Exception as e:
        logger.error(f"Start test error: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

@require_POST
def rmib_autosave_api(request, student_id):
    try:
        student = Student.objects.get(pk=student_id)
        data = json.loads(request.body)
        levels = data.get('levels', {})
        
        # Get or create RMIB result
        rmib_result, created = RMIBResult.objects.get_or_create(
            student=student,
            defaults={'levels': levels, 'status': 'in_progress'}
        )
        
        if not created:
            # Merge with existing
            existing_levels = rmib_result.levels or {}
            existing_levels.update(levels)
            rmib_result.levels = existing_levels
            rmib_result.status = 'in_progress'
        
        # Calculate scores
        rmib_result.calculate_scores()
        rmib_result.save()
        
        return JsonResponse({
            'success': True,
            'scores': rmib_result.scores,
            'message': 'Progress saved'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
# ==================== EXPORT VIEWS - ENHANCED WITH PASSWORD ====================
@login_required
def export_page(request):
    """Display export options page"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki izin untuk mengakses halaman ini')
        return redirect('students:list')
    
    # Get filter stats for UI
    students = Student.objects.all()
    context = {
        'page_title': 'Export Data Siswa',
        'breadcrumb_title': 'Export',
        'total_students': students.count(),
        'available_classes': students.values_list('student_class', flat=True).distinct().order_by('student_class'),
        'available_years': students.values_list('entry_year', flat=True).distinct().order_by('-entry_year'),
        'status_choices': Student.STATUS_CHOICES,
        'gender_choices': Student.GENDER_CHOICES,
    }
    return render(request, 'students/export_page.html', context)


@login_required
def export_students_csv(request):
    """Enhanced CSV export with filters and PASSWORD"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki izin untuk mengekspor data')
        return redirect('students:list')
    
    # Apply filters from GET parameters
    students = Student.objects.select_related('user').all()
    
    # Filter by search
    search = request.GET.get('search', '').strip()
    if search:
        students = students.filter(
            Q(name__icontains=search) |
            Q(nisn__icontains=search) |
            Q(student_class__icontains=search)
        )
    
    # Filter by class
    class_filter = request.GET.get('class', '').strip()
    if class_filter:
        students = students.filter(student_class=class_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '').strip()
    if status_filter:
        students = students.filter(test_status=status_filter)
    
    # Filter by gender
    gender_filter = request.GET.get('gender', '').strip()
    if gender_filter:
        students = students.filter(gender=gender_filter)
    
    # Filter by year
    year_filter = request.GET.get('year', '').strip()
    if year_filter:
        students = students.filter(entry_year=int(year_filter))
    
    # Order by
    students = students.order_by('student_class', 'name')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"data_siswa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Headers - INCLUDE PASSWORD
    writer.writerow([
        'NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 
        'Tanggal Lahir', 'Tempat Lahir', 'Password',
        'Status Tes', 'Tahun Masuk', 'Telepon', 'Alamat', 'Telepon Orang Tua'
    ])
    
    # Data rows
    for student in students:
        writer.writerow([
            student.nisn,
            student.name,
            student.student_class,
            student.get_gender_display(),
            student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '',
            student.birth_place,
            student.generated_password or 'N/A',  # Include password
            student.get_test_status_display(),
            student.entry_year,
            student.phone or '',
            student.address or '',
            student.parent_phone or '',
        ])
    
    logger.info(f'CSV exported WITH PASSWORD: {students.count()} students by {request.user.username}')
    return response


@login_required
def export_students_excel(request):
    """Export to Excel with formatting and PASSWORD - requires openpyxl"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Library openpyxl tidak terinstall. Gunakan CSV export.')
        return redirect('students:export_page')
    
    # Apply same filters as CSV
    students = Student.objects.select_related('user').all()
    
    search = request.GET.get('search', '').strip()
    if search:
        students = students.filter(
            Q(name__icontains=search) |
            Q(nisn__icontains=search) |
            Q(student_class__icontains=search)
        )
    
    class_filter = request.GET.get('class', '').strip()
    if class_filter:
        students = students.filter(student_class=class_filter)
    
    status_filter = request.GET.get('status', '').strip()
    if status_filter:
        students = students.filter(test_status=status_filter)
    
    gender_filter = request.GET.get('gender', '').strip()
    if gender_filter:
        students = students.filter(gender=gender_filter)
    
    year_filter = request.GET.get('year', '').strip()
    if year_filter:
        students = students.filter(entry_year=int(year_filter))
    
    students = students.order_by('student_class', 'name')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Siswa"
    
    # Styling
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    password_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Yellow for password
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - INCLUDE PASSWORD
    headers = [
        'NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 
        'Tanggal Lahir', 'Tempat Lahir', 'Password',
        'Status Tes', 'Tahun Masuk', 'Telepon', 'Alamat', 'Telepon Orang Tua'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, student in enumerate(students, 2):
        data = [
            student.nisn,
            student.name,
            student.student_class,
            student.get_gender_display(),
            student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '',
            student.birth_place,
            student.generated_password or 'N/A',  # Include password
            student.get_test_status_display(),
            student.entry_year,
            student.phone or '',
            student.address or '',
            student.parent_phone or '',
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Highlight password column
            if col_num == 7:  # Password column
                cell.fill = password_fill
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"data_siswa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    logger.info(f'Excel exported WITH PASSWORD: {students.count()} students by {request.user.username}')
    return response


@login_required
def export_rmib_results_csv(request):
    """Export RMIB results to CSV with LEVELS - FIXED"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    # Get only students with completed RMIB
    students = Student.objects.filter(
        test_status='completed'
    ).select_related('rmib_result').order_by('student_class', 'name')
    
    # Apply filters
    class_filter = request.GET.get('class', '').strip()
    if class_filter:
        students = students.filter(student_class=class_filter)
    
    year_filter = request.GET.get('year', '').strip()
    if year_filter:
        students = students.filter(entry_year=int(year_filter))
    
    # Create CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    filename = f"hasil_rmib_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Headers - matching RMIB batch import format
    writer.writerow([
        'NISN', 'Nama', 'Kelas', 'Jenis Kelamin', 'Tanggal Lahir',
        'Out', 'Me', 'COMP', 'Sci', 'Prs', 'Aesth', 
        'Lit', 'Mus', 'S.S', 'Cler', 'Prac', 'Med'
    ])
    
    # RMIB category mapping (levels to CSV columns)
    category_mapping = {
        'outdoor': 'Out',
        'mechanical': 'Me',
        'computational': 'COMP',
        'scientific': 'Sci',
        'personal_contact': 'Prs',
        'aesthetic': 'Aesth',
        'literary': 'Lit',
        'musical': 'Mus',
        'social_service': 'S.S',
        'clerical': 'Cler',
        'practical': 'Prac',
        'medical': 'Med'
    }
    
    exported_count = 0
    
    # Data rows
    for student in students:
        if hasattr(student, 'rmib_result') and student.rmib_result.levels:
            levels = student.rmib_result.levels
            
            # Convert levels dict to ordered list matching CSV headers
            rmib_values = []
            for key in category_mapping.keys():
                level = levels.get(key, 0)
                rmib_values.append(level)
            
            writer.writerow([
                student.nisn,
                student.name,
                student.student_class,
                student.gender,  # L or P
                student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '',
                *rmib_values  # Unpack the 12 RMIB values
            ])
            exported_count += 1
    
    logger.info(f'RMIB CSV exported: {exported_count} results by {request.user.username}')
    return response


@login_required  
def get_export_preview(request):
    """AJAX endpoint to get preview of data to be exported"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        # Apply same filters
        students = Student.objects.all()
        
        search = request.GET.get('search', '').strip()
        if search:
            students = students.filter(
                Q(name__icontains=search) |
                Q(nisn__icontains=search) |
                Q(student_class__icontains=search)
            )
        
        class_filter = request.GET.get('class', '')
        if class_filter:
            students = students.filter(student_class=class_filter)
        
        status_filter = request.GET.get('status', '')
        if status_filter:
            students = students.filter(test_status=status_filter)
        
        gender_filter = request.GET.get('gender', '')
        if gender_filter:
            students = students.filter(gender=gender_filter)
        
        year_filter = request.GET.get('year', '')
        if year_filter:
            students = students.filter(entry_year=int(year_filter))
        
        count = students.count()
        
        # Sample preview (first 5)
        preview_data = []
        for student in students[:5]:
            preview_data.append({
                'nisn': student.nisn,
                'name': student.name,
                'class': student.student_class,
                'gender': student.get_gender_display(),
                'status': student.get_test_status_display(),
            })
        
        return JsonResponse({
            'success': True,
            'count': count,
            'preview': preview_data,
            'has_more': count > 5
        })
    
    except Exception as e:
        logger.error(f'Export preview error: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
